import pandas as pd
import mysql.connector
import json
from datetime import datetime, date, time
import ctypes, sys
import os.path
import openpyxl
import aspose.words as aw
from PyQt5 import QtCore, QtGui, QtWidgets
from aspose.cells import Workbook, SaveFormat, FileFormatType
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, \
    QPushButton, QMainWindow, QLabel
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import (QLabel, QPushButton, QPlainTextEdit, QApplication, QCheckBox, QMainWindow, QWidget,
                             QVBoxLayout, QTabWidget)
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtWidgets import *
from PyQt5.Qt import *

with open('config_path.json', encoding="utf8") as conf:
    config = json.load(conf)

with open(config[0]['config_db']+'/config_db.json', encoding="utf8") as save:
    json_db = json.load(save)

# Подключение к БД
conn = mysql.connector.connect(user=json_db[0]['login'], password=json_db[0]['password'], host=json_db[0]['host'],
                                   database=json_db[0]['name_db'])
cursor = conn.cursor(buffered=True)


class Ui_DeleteStudent(QtWidgets.QWidget):
    def setupUi(self, DeleteStudent):
        DeleteStudent.setObjectName("DeleteStudent")
        DeleteStudent.resize(496, 265)
        self.delete_btn = QtWidgets.QPushButton(DeleteStudent)
        self.delete_btn.setGeometry(QRect(230, 200, 51, 23))
        self.delete_btn.setObjectName("DeleteStudent")
        self.retranslateUi(DeleteStudent)
        QtCore.QMetaObject.connectSlotsByName(DeleteStudent)

    def retranslateUi(self, DeleteStudent):
        _translate = QCoreApplication.translate
        DeleteStudent.setWindowTitle(_translate("DeleteStudent", "Удаление"))
        self.delete_btn.setText(_translate("DeleteStudent", "Удалить"))


class DeleteStudent(QtWidgets.QDialog, Ui_DeleteStudent):
    def __init__(self, data, parent=None):
        gr_name = []
        cour_name = []
        ye_name = []
        org_name = []
        gr_name.append(data[0])
        cour_name.append(data[1])
        ye_name.append(data[2])
        org_name.append(data[3])
        id_data = []
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        sel_main_id = 'SELECT `id` FROM `students` WHERE `groups_id` = %s AND `courses_id` = %s AND `year_enter_id` + %s ' \
                      'AND `organization_id` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(DeleteStudent, self).__init__(parent)
        self.setupUi(self)
        self.delete_btn.clicked.connect(self.delete_stud)
        stud_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_stud = 'SELECT * FROM `students` WHERE `id` = %s'
            cursor.execute(sel_stud, id_)
            stud_id.append(cursor.fetchone())
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(stud_id[0])+1)
        self.table.setRowCount(len(stud_id))
        self.table.setHorizontalHeaderLabels(
            ["Статус", "Номер", "ФИО", "Номер зачётной книжки", "Номер группы", "Группа", "Курс", "Год поступления", "Организация"])
        self.table.horizontalHeaderItem(1).setToolTip("Column 1")
        self.table.horizontalHeaderItem(2).setToolTip("Column 2")
        self.table.horizontalHeaderItem(3).setToolTip("Column 3")
        self.table.horizontalHeaderItem(4).setToolTip("Column 4")
        self.table.horizontalHeaderItem(5).setToolTip("Column 5")
        self.table.horizontalHeaderItem(6).setToolTip("Column 6")
        self.table.horizontalHeaderItem(7).setToolTip("Column 7")
        self.table.horizontalHeaderItem(8).setToolTip("Column 8")
        self.table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(5).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(6).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(7).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(8).setTextAlignment(Qt.AlignHCenter)
        for i in range(0, len(stud_id)):
            sel_stud_name = 'SELECT ' \
                            '(SELECT `id` FROM `students` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `students` WHERE `name` = %s), ' \
                            '(SELECT `number_stud_tiket` FROM `students` WHERE `number_stud_tiket` = %s), ' \
                            '(SELECT `num_group` FROM `students` WHERE `num_group` = %s LIMIT 1), ' \
                            '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `organization` WHERE `id` = %s) ' \
                            'FROM `students`'
            cursor.execute(sel_stud_name, stud_id[i])
            st = cursor.fetchone()
            for j in range(0, len(st)):
                item = QTableWidgetItem(str(st[j]))
                item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(i, j + 1, item)
            widget = QWidget()
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            layoutH = QHBoxLayout(widget)
            layoutH.addWidget(checkbox)
            layoutH.setAlignment(Qt.AlignCenter)
            layoutH.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(i, 0, widget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.delete_btn)

    def delete_stud(self):
        checked_list = []
        for i in range(self.table.rowCount()):
            if self.table.cellWidget(i, 0).findChild(type(QCheckBox())).isChecked():
                checked_list.append(self.table.item(i, 1).text())
        for i in range(0, len(checked_list)):
            id_ = []
            id_.append(checked_list[i])
            del_vopr = 'DELETE FROM `students` WHERE `id` = %s'
            cursor.execute(del_vopr, id_)
            conn.commit()


class Ui_UpdateStudents(QtWidgets.QWidget):
    def setupUi(self, UpdateStudents):
        UpdateStudents.setObjectName("UpdateStudents")
        UpdateStudents.resize(496, 265)
        self.update_btn = QtWidgets.QPushButton(UpdateStudents)
        self.update_btn.setGeometry(QRect(230, 200, 51, 23))
        self.update_btn.setObjectName("UpdateStudents")
        self.retranslateUi(UpdateStudents)
        QtCore.QMetaObject.connectSlotsByName(UpdateStudents)

    def retranslateUi(self, UpdateStudents):
        _translate = QCoreApplication.translate
        UpdateStudents.setWindowTitle(_translate("UpdateStudents", "Редактирование"))
        self.update_btn.setText(_translate("UpdateStudents", "Редактировать"))


class UpdateStudents(QtWidgets.QDialog, Ui_UpdateStudents):
    def __init__(self, data, parent=None):
        gr_name = []
        cour_name = []
        ye_name = []
        org_name = []
        gr_name.append(data[0])
        cour_name.append(data[1])
        ye_name.append(data[2])
        org_name.append(data[3])
        id_data = []
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        sel_main_id = 'SELECT `id` FROM `students` WHERE `groups_id` = %s AND `courses_id` = %s AND `year_enter_id` + %s ' \
                      'AND `organization_id` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(UpdateStudents, self).__init__(parent)
        self.setupUi(self)
        self.update_btn.clicked.connect(self.update_stud)
        stud_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_stud = 'SELECT * FROM `students` WHERE `id` = %s'
            cursor.execute(sel_stud, id_)
            stud_id.append(cursor.fetchone())
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(stud_id[0]))
        self.table.setRowCount(len(stud_id))
        self.table.setHorizontalHeaderLabels(
            ["Номер", "ФИО", "Номер зачётной книжки", "Номер группы", "Группа", "Курс", "Год поступления", "Организация"])
        print(stud_id)
        for i in range(0, len(stud_id)):
            sel_stud_name = 'SELECT ' \
                            '(SELECT `id` FROM `students` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `students` WHERE `name` = %s), ' \
                            '(SELECT `number_stud_tiket` FROM `students` WHERE `number_stud_tiket` = %s), ' \
                            '(SELECT `num_group` FROM `students` WHERE `num_group` = %s LIMIT 1), ' \
                            '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `organization` WHERE `id` = %s) ' \
                            'FROM `students`'
            cursor.execute(sel_stud_name, stud_id[i])
            st = cursor.fetchone()
            for j in range(0, len(st)):
                if j == 0:
                    item = QTableWidgetItem(str(st[j]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table.setItem(i, j, item)
                else:
                    self.table.setItem(i, j, QTableWidgetItem(str(st[j])))
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.update_btn)

    def update_stud(self):
        for i in range(self.table.rowCount()):
            tb_data = []
            for j in range(0, self.table.columnCount()):
                tb_data.append(self.table.item(i, j).text())
            check_gr = 'SELECT * FROM `groups` WHERE `name` = %s'
            gr_name = []
            gr_name.append(tb_data[4])
            cursor.execute(check_gr, gr_name)
            gr_id = cursor.fetchone()
            if gr_id == None:
                in_gr = 'INSERT INTO `groups` (`name`) VALUES (%s)'
                cursor.execute(in_gr, gr_name)
                conn.commit()
            check_cour = 'SELECT * FROM `courses` WHERE `name` = %s'
            cour_name = []
            cour_name.append(tb_data[5])
            cursor.execute(check_cour, cour_name)
            cour_id = cursor.fetchone()
            if cour_id == None:
                in_cour = 'INSERT INTO `courses` (`name`) VALUES (%s)'
                cursor.execute(in_cour, cour_name)
                conn.commit()
            check_ye = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            ye_name = []
            ye_name.append(tb_data[6])
            cursor.execute(check_ye, ye_name)
            ye_id = cursor.fetchone()
            if ye_id == None:
                in_ye = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
                cursor.execute(in_ye, ye_name)
                conn.commit()
            check_org = 'SELECT * FROM `organization` WHERE `name` = %s'
            org_name = []
            org_name.append(tb_data[7])
            cursor.execute(check_org, org_name)
            org_id = cursor.fetchone()
            if org_id == None:
                in_org = 'INSERT INTO `organization` (`name`) VALUES (%s)'
                cursor.execute(in_org, org_name)
                conn.commit()
            id_tb_data = []
            id_tb_data.append(tb_data[0])
            id_tb_data.append(tb_data[4])
            id_tb_data.append(tb_data[5])
            id_tb_data.append(tb_data[6])
            id_tb_data.append(tb_data[7])
            sel_id = 'SELECT ' \
                    '(SELECT `id` FROM `students` WHERE `id` = %s), ' \
                    '(SELECT `id` FROM `groups` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `courses` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `year_enter` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `organization` WHERE `name` = %s) ' \
                    'FROM `students`'
            cursor.execute(sel_id, id_tb_data)
            id_data = list(cursor.fetchone())
            id_data.append(tb_data[1])
            id_data.append(tb_data[2])
            id_data.append(tb_data[3])
            check_db = 'SELECT * FROM `students` WHERE `id` = %s AND `groups_id` = %s AND `courses_id` = %s AND ' \
                       '`year_enter_id` = %s AND `organization_id` = %s AND `name` = %s AND `number_stud_tiket` = %s ' \
                       'AND `num_group` = %s'
            cursor.execute(check_db, id_data)
            check_ = cursor.fetchone()
            if check_ == None:
                up_id = []
                up_id.append(id_data[1])
                up_id.append(id_data[2])
                up_id.append(id_data[3])
                up_id.append(id_data[4])
                up_id.append(id_data[5])
                up_id.append(id_data[6])
                up_id.append(id_data[7])
                up_id.append(id_data[0])
                up_t = 'UPDATE `students` SET `groups_id` = %s, `courses_id` = %s, `year_enter_id` = %s, ' \
                       '`organization_id` = %s, `name` = %s, `number_stud_tiket` = %s, `num_group` = %s WHERE `id` = %s'
                cursor.execute(up_t, up_id)
                conn.commit()


class Ui_UpdateWeekDate(QtWidgets.QWidget):
    def setupUi(self, UpdateWeekDate):
        UpdateWeekDate.setObjectName("UpdateWeekDate")
        UpdateWeekDate.resize(496, 265)
        self.update_btn = QtWidgets.QPushButton(UpdateWeekDate)
        self.update_btn.setGeometry(QRect(230, 200, 51, 23))
        self.update_btn.setObjectName("UpdateWeekDate")
        self.retranslateUi(UpdateWeekDate)
        QtCore.QMetaObject.connectSlotsByName(UpdateWeekDate)

    def retranslateUi(self, UpdateWeekDate):
        _translate = QCoreApplication.translate
        UpdateWeekDate.setWindowTitle(_translate("UpdateWeekDate", "Редактирование"))
        self.update_btn.setText(_translate("UpdateWeekDate", "Редактировать"))


class UpdateWeekDate(QtWidgets.QDialog, Ui_UpdateWeekDate):
    def __init__(self, data, parent=None):
        tt_name = []
        tt_name.append(data[0])
        id_data = []
        sel_sub = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        cursor.execute(sel_sub, tt_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(data[1])
        sel_main_id = 'SELECT `id` FROM `date_type_week` WHERE `type_week_id` = %s AND `date_week` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(UpdateWeekDate, self).__init__(parent)
        self.setupUi(self)
        self.update_btn.clicked.connect(self.update_week_date)
        wd_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_wd = 'SELECT * FROM `date_type_week` WHERE `id` = %s'
            cursor.execute(sel_wd, id_)
            wd_id.append(cursor.fetchone())
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(wd_id[0]))
        self.table.setRowCount(len(wd_id))
        self.table.setHorizontalHeaderLabels(
            ["Номер", "Тип недели", "Дата"])
        for i in range(0, len(wd_id)):
            sel_wd_name = 'SELECT ' \
                            '(SELECT `id` FROM `date_type_week` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `type_week` WHERE `id` = %s), ' \
                            '(SELECT `date_week` FROM `date_type_week` WHERE `date_week` = %s LIMIT 1) ' \
                            'FROM `date_type_week`'
            cursor.execute(sel_wd_name, wd_id[i])
            wd = cursor.fetchone()
            for j in range(0, len(wd)):
                if j == 0:
                    item = QTableWidgetItem(str(wd[j]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table.setItem(i, j, item)
                else:
                    self.table.setItem(i, j, QTableWidgetItem(str(wd[j])))
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.update_btn)

    def update_week_date(self):
        for i in range(self.table.rowCount()):
            tb_data = []
            for j in range(0, self.table.columnCount()):
                tb_data.append(self.table.item(i, j).text())
            check_sub = 'SELECT * FROM `type_week` WHERE `name` = %s'
            tt_name = []
            tt_name.append(tb_data[1])
            cursor.execute(check_sub, tt_name)
            sub_id = cursor.fetchone()
            if sub_id == None:
                in_sub = 'INSERT INTO `type_week` (`name`) VALUES (%s)'
                cursor.execute(in_sub, tt_name)
                conn.commit()
            id_tb_data = []
            id_tb_data.append(tb_data[0])
            id_tb_data.append(tb_data[1])
            sel_id = 'SELECT ' \
                    '(SELECT `id` FROM `date_type_week` WHERE `id` = %s), ' \
                    '(SELECT `id` FROM `type_week` WHERE `name` = %s) ' \
                    'FROM `lessons_plan`'
            cursor.execute(sel_id, id_tb_data)
            id_data = list(cursor.fetchone())
            id_data.append(tb_data[2])
            check_db = 'SELECT * FROM `date_type_week` WHERE `id` = %s AND `type_week_id` = %s AND ' \
                      '`date_week` = %s'
            cursor.execute(check_db, id_data)
            check_ = cursor.fetchone()
            if check_ == None:
                up_id = []
                up_id.append(id_data[1])
                up_id.append(id_data[2])
                up_id.append(id_data[0])
                up_t = 'UPDATE `date_type_week` SET `type_week_id` = %s, `date_week` = %s WHERE `id` = %s'
                cursor.execute(up_t, up_id)
                conn.commit()


class Ui_UpdateTheme(QtWidgets.QWidget):
    def setupUi(self, UpdateTheme):
        UpdateTheme.setObjectName("UpdateTheme")
        UpdateTheme.resize(496, 265)
        self.update_btn = QtWidgets.QPushButton(UpdateTheme)
        self.update_btn.setGeometry(QRect(230, 200, 51, 23))
        self.update_btn.setObjectName("update_btn")
        self.retranslateUi(UpdateTheme)
        QtCore.QMetaObject.connectSlotsByName(UpdateTheme)

    def retranslateUi(self, UpdateTheme):
        _translate = QCoreApplication.translate
        UpdateTheme.setWindowTitle(_translate("UpdateTheme", "Редактирование"))
        self.update_btn.setText(_translate("UpdateTheme", "Редактировать"))


class UpdateTheme(QtWidgets.QDialog, Ui_UpdateTheme):
    def __init__(self, data, parent=None):
        sub_name = []
        gr_name = []
        cour_name = []
        ye_name = []
        org_name = []
        sub_name.append(data[0])
        gr_name.append(data[1])
        cour_name.append(data[2])
        ye_name.append(data[3])
        org_name.append(data[4])
        id_data = []
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        cursor.execute(sel_sub, sub_name)
        id_data.append(cursor.fetchone()[0])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[5]))
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[6]))
        sel_main_id = 'SELECT `id` FROM `lessons_plan` WHERE `subjects_id` = %s AND `groups_id` = %s AND `courses_id` ' \
                      '= %s AND `year_enter_id` = %s AND `number` = %s AND `organization_id` = %s AND `term` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(UpdateTheme, self).__init__(parent)
        self.setupUi(self)
        self.update_btn.clicked.connect(self.update_theme_hours)
        th_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_th = 'SELECT * FROM `lessons_plan` WHERE `id` = %s'
            cursor.execute(sel_th, id_)
            th_id.append(cursor.fetchone())
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(th_id[0]))
        self.table.setRowCount(len(th_id))
        self.table.setHorizontalHeaderLabels(
            ["Номер", "Дисциплина", "Тема", "Группа", "Курсы", "Год поступления", "Количество часов",
             "Организация", "Семестр"])
        for i in range(0, len(th_id)):
            sel_th_name = 'SELECT ' \
                            '(SELECT `id` FROM `lessons_plan` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                            '(SELECT `theme` FROM `lessons_plan` WHERE `theme` = %s LIMIT 1), ' \
                            '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                            '(SELECT `number` FROM `lessons_plan` WHERE `number` = %s LIMIT 1), ' \
                            '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                            '(SELECT `term` FROM `lessons_plan` WHERE `term` = %s LIMIT 1) ' \
                            'FROM `lessons_plan`'
            cursor.execute(sel_th_name, th_id[i])
            th = cursor.fetchone()
            for j in range(0, len(th)):
                if j == 0:
                    item = QTableWidgetItem(str(th[j]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table.setItem(i, j, item)
                else:
                    self.table.setItem(i, j, QTableWidgetItem(str(th[j])))
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.update_btn)

    def update_theme_hours(self):
        for i in range(self.table.rowCount()):
            tb_data = []
            for j in range(0, self.table.columnCount()):
                tb_data.append(self.table.item(i, j).text())
            check_sub = 'SELECT * FROM `subjects` WHERE `name` = %s'
            sub_name = []
            sub_name.append(tb_data[1])
            cursor.execute(check_sub, sub_name)
            sub_id = cursor.fetchone()
            if sub_id == None:
                in_sub = 'INSERT INTO `subjects` (`name`) VALUES (%s)'
                cursor.execute(in_sub, sub_name)
                conn.commit()
            check_gr = 'SELECT * FROM `groups` WHERE `name` = %s'
            gr_name = []
            gr_name.append(tb_data[3])
            cursor.execute(check_gr, gr_name)
            gr_id = cursor.fetchone()
            if gr_id == None:
                in_gr = 'INSERT INTO `groups` (`name`) VALUES (%s)'
                cursor.execute(in_gr, gr_name)
                conn.commit()
            check_cour = 'SELECT * FROM `courses` WHERE `name` = %s'
            cour_name = []
            cour_name.append(tb_data[4])
            cursor.execute(check_cour, cour_name)
            cour_id = cursor.fetchone()
            if cour_id == None:
                in_cour = 'INSERT INTO `courses` (`name`) VALUES (%s)'
                cursor.execute(in_cour, cour_name)
                conn.commit()
            check_ye = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            ye_name = []
            ye_name.append(tb_data[5])
            cursor.execute(check_ye, ye_name)
            ye_id = cursor.fetchone()
            if ye_id == None:
                in_ye = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
                cursor.execute(in_ye, ye_name)
                conn.commit()
            check_org = 'SELECT * FROM `organization` WHERE `name` = %s'
            org_name = []
            org_name.append(tb_data[7])
            cursor.execute(check_org, org_name)
            org_id = cursor.fetchone()
            if org_id == None:
                in_org = 'INSERT INTO `organization` (`name`) VALUES (%s)'
                cursor.execute(in_org, org_name)
                conn.commit()
            id_tb_data = []
            id_tb_data.append(tb_data[0])
            id_tb_data.append(tb_data[1])
            id_tb_data.append(tb_data[3])
            id_tb_data.append(tb_data[4])
            id_tb_data.append(tb_data[5])
            id_tb_data.append(tb_data[7])
            sel_id = 'SELECT ' \
                    '(SELECT `id` FROM `lessons_plan` WHERE `id` = %s), ' \
                    '(SELECT `id` FROM `subjects` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `groups` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `courses` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `year_enter` WHERE `name` = %s), ' \
                    '(SELECT `id` FROM `organization` WHERE `name` = %s) ' \
                    'FROM `lessons_plan`'
            cursor.execute(sel_id, id_tb_data)
            id_data = list(cursor.fetchone())
            id_data.append(tb_data[2])
            id_data.append(tb_data[6])
            id_data.append(tb_data[8])
            check_db = 'SELECT * FROM `lessons_plan` WHERE `id` = %s AND `subjects_id` = %s AND ' \
                      '`groups_id` = %s AND `courses_id` = %s AND `year_enter_id` = %s AND ' \
                      '`organization_id` = %s AND `theme` = %s AND `number` = %s AND `term` = %s'
            cursor.execute(check_db, id_data)
            check_ = cursor.fetchone()
            if check_ == None:
                up_id = []
                up_id.append(id_data[1])
                up_id.append(id_data[2])
                up_id.append(id_data[3])
                up_id.append(id_data[4])
                up_id.append(id_data[5])
                up_id.append(id_data[6])
                up_id.append(id_data[7])
                up_id.append(id_data[8])
                up_id.append(id_data[0])
                up_t = 'UPDATE `lessons_plan` SET `subjects_id` = %s, `groups_id` = %s, `courses_id` = %s, ' \
                       '`year_enter_id` = %s, `organization_id` = %s, `theme` = %s, `number` = %s, `term` = %s ' \
                       'WHERE `id` = %s'
                cursor.execute(up_t, up_id)
                conn.commit()


class Ui_UpdateScheduleChanges(QtWidgets.QWidget):
    def setupUi(self, UpdateScheduleChanges):
        UpdateScheduleChanges.setObjectName("UpdateScheduleChanges")
        UpdateScheduleChanges.resize(496, 265)
        self.update_btn = QtWidgets.QPushButton(UpdateScheduleChanges)
        self.update_btn.setGeometry(QRect(230, 200, 51, 23))
        self.update_btn.setObjectName("update_btn")
        self.retranslateUi(UpdateScheduleChanges)
        QtCore.QMetaObject.connectSlotsByName(UpdateScheduleChanges)

    def retranslateUi(self, UpdateScheduleChanges):
        _translate = QCoreApplication.translate
        UpdateScheduleChanges.setWindowTitle(_translate("UpdateScheduleChanges", "Редактирование"))
        self.update_btn.setText(_translate("UpdateScheduleChanges", "Редактировать"))


class UpdateScheduleChanges(QtWidgets.QDialog, Ui_UpdateScheduleChanges):
    def __init__(self, data, parent=None):
        gr_name = []
        cour_name = []
        ye_name = []
        nl_name = []
        sub_name = []
        tea_name = []
        org_name = []
        gr_name.append(data[0])
        cour_name.append(data[1])
        ye_name.append(data[2])
        nl_name.append(data[3])
        sub_name.append(data[4])
        tea_name.append(data[5])
        org_name.append(data[7])
        id_data = []
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_nl = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        cursor.execute(sel_nl, nl_name)
        id_data.append(cursor.fetchone()[0])
        if sub_name[0] == '<<Не определено>>' and tea_name[0] == '<<Не определено>>':
            id_data.append(None)
            id_data.append(None)
            sel_main_id = 'SELECT `id` FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                          '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` is %s AND `teachers_id` is %s AND ' \
                          '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
        else:
            sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
            cursor.execute(sel_sub, sub_name)
            id_data.append(cursor.fetchone()[0])
            sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
            cursor.execute(sel_tea, tea_name)
            id_data.append(cursor.fetchone()[0])
            sel_main_id = 'SELECT `id` FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                          '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` = %s AND `teachers_id` = %s AND ' \
                          '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
        id_data.append(data[6])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[8]))
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(UpdateScheduleChanges, self).__init__(parent)
        self.setupUi(self)
        self.update_btn.clicked.connect(self.update_schedule_changes)
        sch_c_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch = 'SELECT * FROM `schedule_changes` WHERE `id` = %s'
            cursor.execute(sel_sch, id_)
            sch_c_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_c_id[0]))
        self.table.setRowCount(len(sch_c_id))
        self.table.setHorizontalHeaderLabels(
            ["Номер", "Группа", "Курсы", "Год поступления", "Номер пары", "Дисциплина", "Преподаватель",  "Дата",
             "Организация", "Номер группы"])
        for i in range(0, len(sch_c_id)):
            sc_id = []
            sc_id.append(sch_c_id[i][0])
            sc_id.append(sch_c_id[i][1])
            sc_id.append(sch_c_id[i][2])
            sc_id.append(sch_c_id[i][3])
            sc_id.append(sch_c_id[i][4])
            sc_id.append(sch_c_id[i][5])
            sc_id.append(sch_c_id[i][6])
            sc_id.append(sch_c_id[i][7])
            sc_id.append(sch_c_id[i][8])
            sc_id.append(sch_c_id[i][9])
            sel_sch_c_name = 'SELECT ' \
                            '(SELECT `id` FROM `schedule_changes` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `num_lessons` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                            '(SELECT `name` FROM `teachers` WHERE `id` = %s), ' \
                            '(SELECT `date_changes` FROM `schedule_changes` WHERE `date_changes` = %s LIMIT 1), ' \
                            '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                            '(SELECT `num_group` FROM `schedule_changes` WHERE `num_group` = %s LIMIT 1) ' \
                            'FROM `schedule_changes`'
            cursor.execute(sel_sch_c_name, sc_id)
            sch_c = cursor.fetchone()
            for j in range(0, len(sch_c)):
                if j == 0:
                    item = QTableWidgetItem(str(sch_c[j]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table.setItem(i, j, item)
                else:
                    self.table.setItem(i, j, QTableWidgetItem(str(sch_c[j])))
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.update_btn)

    def update_schedule_changes(self):
        for i in range(self.table.rowCount()):
            tb_data = []
            for j in range(0, self.table.columnCount()):
                tb_data.append(self.table.item(i, j).text())
            check_gr = 'SELECT * FROM `groups` WHERE `name` = %s'
            gr_name = []
            gr_name.append(tb_data[1])
            cursor.execute(check_gr, gr_name)
            gr_id = cursor.fetchone()
            if gr_id == None:
                in_gr = 'INSERT INTO `groups` (`name`) VALUES (%s)'
                cursor.execute(in_gr, gr_name)
                conn.commit()
            check_cour = 'SELECT * FROM `courses` WHERE `name` = %s'
            cour_name = []
            cour_name.append(tb_data[2])
            cursor.execute(check_cour, cour_name)
            cour_id = cursor.fetchone()
            if cour_id == None:
                in_cour = 'INSERT INTO `courses` (`name`) VALUES (%s)'
                cursor.execute(in_cour, cour_name)
                conn.commit()
            check_ye = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            ye_name = []
            ye_name.append(tb_data[3])
            cursor.execute(check_ye, ye_name)
            ye_id = cursor.fetchone()
            if ye_id == None:
                in_ye = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
                cursor.execute(in_ye, ye_name)
                conn.commit()
            check_nl = 'SELECT * FROM `num_lessons` WHERE `name` = %s'
            nl_name = []
            nl_name.append(tb_data[4])
            cursor.execute(check_nl, nl_name)
            nl_id = cursor.fetchone()
            if nl_id == None:
                in_nl = 'INSERT INTO `num_lessons` (`name`) VALUES (%s)'
                cursor.execute(in_nl, nl_name)
                conn.commit()
            check_sub = 'SELECT * FROM `subjects` WHERE `name` = %s'
            sub_name = []
            sub_name.append(tb_data[5])
            cursor.execute(check_sub, sub_name)
            sub_id = cursor.fetchone()
            if sub_id == None:
                in_sub = 'INSERT INTO `subjects` (`name`) VALUES (%s)'
                cursor.execute(in_sub, sub_name)
                conn.commit()
            check_tas = 'SELECT * FROM `teachers` WHERE `name` = %s'
            tea_name = []
            tea_name.append(tb_data[6])
            cursor.execute(check_tas, tea_name)
            tea_id = cursor.fetchone()
            if tea_id == None:
                in_tea = 'INSERT INTO `teachers` (`name`) VALUES (%s)'
                cursor.execute(in_tea, tea_name)
                conn.commit()
            check_org = 'SELECT * FROM `organization` WHERE `name` = %s'
            org_name = []
            org_name.append(tb_data[8])
            cursor.execute(check_org, org_name)
            org_id = cursor.fetchone()
            if org_id == None:
                in_org = 'INSERT INTO `organization` (`name`) VALUES (%s)'
                cursor.execute(in_org, org_name)
                conn.commit()
            if tb_data[5] == None or tb_data[5] == 'None':
                tb_data[5] = None
                tb_data[6] = None
                sel_id = 'SELECT ' \
                        '(SELECT `id` FROM `schedule_changes` WHERE `id` = %s), ' \
                        '(SELECT `id` FROM `groups` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `courses` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `year_enter` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `num_lessons` WHERE `name` = %s), ' \
                        '(SELECT `subjects_id` FROM `schedule_changes` WHERE `subjects_id` is %s), ' \
                        '(SELECT `teachers_id` FROM `schedule_changes` WHERE `subjects_id` is %s), ' \
                        '(SELECT `date_changes` FROM `schedule_changes` WHERE `date_changes` = %s LIMIT 1), ' \
                        '(SELECT `id` FROM `organization` WHERE `name` = %s), ' \
                        '(SELECT `num_group` FROM `schedule_changes` WHERE `num_group` = %s LIMIT 1) ' \
                        'FROM `schedule_changes`'
            else:
                sel_id = 'SELECT ' \
                        '(SELECT `id` FROM `schedule_changes` WHERE `id` = %s), ' \
                        '(SELECT `id` FROM `groups` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `courses` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `year_enter` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `num_lessons` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `subjects` WHERE `name` = %s), ' \
                        '(SELECT `id` FROM `teachers` WHERE `name` = %s), ' \
                        '(SELECT `date_changes` FROM `schedule_changes` WHERE `date_changes` = %s LIMIT 1), ' \
                        '(SELECT `id` FROM `organization` WHERE `name` = %s), ' \
                        '(SELECT `num_group` FROM `schedule_changes` WHERE `num_group` = %s LIMIT 1) ' \
                        'FROM `schedule_changes`'
            cursor.execute(sel_id, tb_data)
            id_data = cursor.fetchone()
            if tb_data[5] == None or tb_data[5] == 'None':
                check_db = 'SELECT * FROM `schedule_changes` WHERE `id` = %s AND `groups_id` = %s AND `courses_id` = %s AND ' \
                           '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` is %s AND `teachers_id` is %s AND ' \
                           '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
            else:
                check_db = 'SELECT * FROM `schedule_changes` WHERE `id` = %s AND `groups_id` = %s AND `courses_id` = %s AND ' \
                          '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` = %s AND `teachers_id` = %s AND ' \
                          '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
            cursor.execute(check_db, id_data)
            check_ = cursor.fetchone()
            if check_ == None:
                up_id = []
                up_id.append(id_data[1])
                up_id.append(id_data[2])
                up_id.append(id_data[3])
                up_id.append(id_data[4])
                up_id.append(id_data[5])
                up_id.append(id_data[6])
                up_id.append(id_data[7])
                up_id.append(id_data[8])
                up_id.append(id_data[9])
                up_id.append(id_data[0])
                up_t = 'UPDATE `schedule_changes` SET `groups_id` = %s, `courses_id` = %s, `year_enter_id` = %s, ' \
                       '`num_lessons_id` = %s, `subjects_id` = %s, `teachers_id` = %s, `date_changes` = %s, ' \
                       '`organization_id` = %s, `num_group` = %s WHERE `id` = %s'
                cursor.execute(up_t, up_id)
                conn.commit()


class Ui_UpdateSchedule(QtWidgets.QWidget):
    def setupUi(self, UpdateSchedule):
        UpdateSchedule.setObjectName("UpdateSchedule")
        UpdateSchedule.resize(496, 265)
        self.update_btn = QtWidgets.QPushButton(UpdateSchedule)
        self.update_btn.setGeometry(QRect(230, 200, 51, 23))
        self.update_btn.setObjectName("update_btn")
        self.retranslateUi(UpdateSchedule)
        QtCore.QMetaObject.connectSlotsByName(UpdateSchedule)

    def retranslateUi(self, UpdateSchedule):
        _translate = QCoreApplication.translate
        UpdateSchedule.setWindowTitle(_translate("UpdateSchedule", "Редактирование"))
        self.update_btn.setText(_translate("UpdateSchedule", "Редактировать"))


class UpdateSchedule(QtWidgets.QDialog, Ui_UpdateSchedule):
    def __init__(self, data, parent=None):
        sub_name = []
        tea_name = []
        gr_name = []
        cour_name = []
        ye_name = []
        nl_name = []
        nd_name = []
        tw_name = []
        org_name = []
        ss_name = []
        sub_name.append(data[0])
        tea_name.append(data[1])
        gr_name.append(data[2])
        cour_name.append(data[3])
        ye_name.append(data[4])
        nl_name.append(data[5])
        nd_name.append(data[6])
        tw_name.append(data[7])
        org_name.append(data[8])
        ss_name.append(data[9])
        id_data = []
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        cursor.execute(sel_sub, sub_name)
        id_data.append(cursor.fetchone()[0])
        sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
        cursor.execute(sel_tea, tea_name)
        id_data.append(cursor.fetchone()[0])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_per = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        cursor.execute(sel_per, nl_name)
        id_data.append(cursor.fetchone()[0])
        sel_tt = 'SELECT `id` FROM `name_day` WHERE `name` = %s'
        cursor.execute(sel_tt, nd_name)
        id_data.append(cursor.fetchone()[0])
        sel_bl = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        cursor.execute(sel_bl, tw_name)
        id_data.append(cursor.fetchone()[0])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        sel_ss = 'SELECT `id` FROM `sprav_schedule` WHERE `name` = %s'
        cursor.execute(sel_ss, ss_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[10]))
        main_id = []
        sel_main_id = 'SELECT `id` FROM `schedule` WHERE `subjects_id` = %s AND `teachers_id` = %s AND `groups_id` = %s AND ' \
                      '`courses_id` = %s AND `year_enter_id` = %s AND `num_lessons_id` = %s AND `name_day_id` = %s AND ' \
                      '`type_week_id` = %s AND `organization_id` = %s AND `sprav_schedule_id` = %s AND `num_group` = %s'
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(UpdateSchedule, self).__init__(parent)
        self.setupUi(self)
        self.update_btn.clicked.connect(self.update_schedule)
        sch_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch = 'SELECT * FROM `schedule` WHERE `id` = %s'
            cursor.execute(sel_sch, id_)
            sch_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_id[0]))
        self.table.setRowCount(len(sch_id))
        self.table.setHorizontalHeaderLabels(
            ["Номер", "Дисциплина", "Преподаватель", "Группа", "Курсы", "Год поступления", "Номер пары", "День недели",
             "Тип недели", "Организация", "Номер группы", "Основное расписание"])
        for i in range(0, len(sch_id)):
            s_id = []
            s_id.append(sch_id[i][0])
            s_id.append(sch_id[i][1])
            s_id.append(sch_id[i][2])
            s_id.append(sch_id[i][3])
            s_id.append(sch_id[i][4])
            s_id.append(sch_id[i][5])
            s_id.append(sch_id[i][6])
            s_id.append(sch_id[i][7])
            s_id.append(sch_id[i][8])
            s_id.append(sch_id[i][9])
            s_id.append(sch_id[i][10])
            s_id.append(sch_id[i][11])
            sel_sch_name = 'SELECT ' \
                           '(SELECT `id` FROM `schedule` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `teachers` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `num_lessons` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `name_day` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `type_week` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                           '(SELECT `num_group` FROM `schedule` WHERE `num_group` = %s LIMIT 1), ' \
                           '(SELECT `name` FROM `sprav_schedule` WHERE `id` = %s) ' \
                           'FROM `schedule`'
            cursor.execute(sel_sch_name, s_id)
            sch = cursor.fetchone()
            for j in range(0, len(sch)):
                if j == 0:
                    item = QTableWidgetItem(str(sch[j]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table.setItem(i, j, item)
                else:
                    self.table.setItem(i, j, QTableWidgetItem(str(sch[j])))
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.update_btn)

    def update_schedule(self):
        for i in range(self.table.rowCount()):
            tb_data = []
            for j in range(0, self.table.columnCount()):
                tb_data.append(self.table.item(i, j).text())
            check_sub = 'SELECT * FROM `subjects` WHERE `name` = %s'
            sub_name = []
            sub_name.append(tb_data[1])
            cursor.execute(check_sub, sub_name)
            sub_id = cursor.fetchone()
            if sub_id == None:
                in_sub = 'INSERT INTO `subjects` (`name`) VALUES (%s)'
                cursor.execute(in_sub, sub_name)
                conn.commit()
            check_tas = 'SELECT * FROM `teachers` WHERE `name` = %s'
            tea_name = []
            tea_name.append(tb_data[2])
            cursor.execute(check_tas, tea_name)
            tea_id = cursor.fetchone()
            if tea_id == None:
                in_tea = 'INSERT INTO `teachers` (`name`) VALUES (%s)'
                cursor.execute(in_tea, tea_name)
                conn.commit()
            check_gr = 'SELECT * FROM `groups` WHERE `name` = %s'
            gr_name = []
            gr_name.append(tb_data[3])
            cursor.execute(check_gr, gr_name)
            gr_id = cursor.fetchone()
            if gr_id == None:
                in_gr = 'INSERT INTO `groups` (`name`) VALUES (%s)'
                cursor.execute(in_gr, gr_name)
                conn.commit()
            check_cour = 'SELECT * FROM `courses` WHERE `name` = %s'
            cour_name = []
            cour_name.append(tb_data[4])
            cursor.execute(check_cour, cour_name)
            cour_id = cursor.fetchone()
            if cour_id == None:
                in_cour = 'INSERT INTO `courses` (`name`) VALUES (%s)'
                cursor.execute(in_cour, cour_name)
                conn.commit()
            check_ye = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            ye_name = []
            ye_name.append(tb_data[5])
            cursor.execute(check_ye, ye_name)
            ye_id = cursor.fetchone()
            if ye_id == None:
                in_ye = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
                cursor.execute(in_ye, ye_name)
                conn.commit()
            check_nl = 'SELECT * FROM `num_lessons` WHERE `name` = %s'
            nl_name = []
            nl_name.append(tb_data[6])
            cursor.execute(check_nl, nl_name)
            nl_id = cursor.fetchone()
            if nl_id == None:
                in_nl = 'INSERT INTO `num_lessons` (`name`) VALUES (%s)'
                cursor.execute(in_nl, nl_name)
                conn.commit()
            check_nd = 'SELECT * FROM `name_day` WHERE `name` = %s'
            nd_name = []
            nd_name.append(tb_data[7])
            cursor.execute(check_nd, nd_name)
            nd_id = cursor.fetchone()
            if nd_id == None:
                in_nd = 'INSERT INTO `name_day` (`name`) VALUES (%s)'
                cursor.execute(in_nd, nd_name)
                conn.commit()
            check_tw = 'SELECT * FROM `type_week` WHERE `name` = %s'
            tw_name = []
            tw_name.append(tb_data[8])
            cursor.execute(check_tw, tw_name)
            tw_id = cursor.fetchone()
            if tw_id == None:
                in_tw = 'INSERT INTO `type_week` (`name`) VALUES (%s)'
                cursor.execute(in_tw, tw_name)
                conn.commit()
            check_org = 'SELECT * FROM `organization` WHERE `name` = %s'
            org_name = []
            org_name.append(tb_data[9])
            cursor.execute(check_org, org_name)
            org_id = cursor.fetchone()
            if org_id == None:
                in_org = 'INSERT INTO `organization` (`name`) VALUES (%s)'
                cursor.execute(in_org, org_name)
                conn.commit()
            check_ss = 'SELECT * FROM `sprav_schedule` WHERE `name` = %s'
            ss_name = []
            ss_name.append(tb_data[11])
            cursor.execute(check_ss, ss_name)
            ss_id = cursor.fetchone()
            if ss_id == None:
                in_ss = 'INSERT INTO `sprav_schedule` (`name`) VALUES (%s)'
                cursor.execute(in_ss, ss_name)
                conn.commit()
            sel_id = 'SELECT ' \
                     '(SELECT `id` FROM `schedule` WHERE `id` = %s), ' \
                     '(SELECT `id` FROM `subjects` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `teachers` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `groups` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `courses` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `year_enter` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `num_lessons` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `name_day` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `type_week` WHERE `name` = %s), ' \
                     '(SELECT `id` FROM `organization` WHERE `name` = %s), ' \
                     '(SELECT `num_group` FROM `schedule` WHERE `num_group` = %s LIMIT 1), ' \
                     '(SELECT `id` FROM `sprav_schedule` WHERE `name` = %s) ' \
                     'FROM `schedule`'
            cursor.execute(sel_id, tb_data)
            id_data = cursor.fetchone()
            check_db = 'SELECT * FROM `schedule` WHERE `id` = %s AND `subjects_id` = %s AND `teachers_id` = %s AND ' \
                       '`groups_id` = %s AND `courses_id` = %s AND `year_enter_id` = %s AND `num_lessons_id` = %s AND ' \
                       '`name_day_id` = %s AND `type_week_id` = %s AND `organization_id` = %s AND `num_group` = %s AND ' \
                       '`sprav_schedule_id` = %s'
            cursor.execute(check_db, id_data)
            check_ = cursor.fetchone()
            if check_ == None:
                up_id = []
                up_id.append(id_data[1])
                up_id.append(id_data[2])
                up_id.append(id_data[3])
                up_id.append(id_data[4])
                up_id.append(id_data[5])
                up_id.append(id_data[6])
                up_id.append(id_data[7])
                up_id.append(id_data[8])
                up_id.append(id_data[9])
                up_id.append(id_data[10])
                up_id.append(id_data[11])
                up_id.append(id_data[0])
                up_t = 'UPDATE `schedule` SET `subjects_id` = %s, `teachers_id` = %s, `groups_id` = %s, ' \
                       '`courses_id` = %s, `year_enter_id` = %s, `num_lessons_id` = %s, `name_day_id` = %s, ' \
                       '`type_week_id` = %s, `organization_id` = %s, `num_group` = %s, `sprav_schedule_id` = %s ' \
                       'WHERE `id` = %s'
                cursor.execute(up_t, up_id)
                conn.commit()


class Ui_DeleteSchedule(QtWidgets.QWidget):
    def setupUi(self, DeleteSchedule):
        DeleteSchedule.setObjectName("DeleteSchedule")
        DeleteSchedule.resize(496, 265)
        self.delete_btn = QtWidgets.QPushButton(DeleteSchedule)
        self.delete_btn.setGeometry(QRect(230, 200, 51, 23))
        self.delete_btn.setObjectName("delete_btn")
        self.retranslateUi(DeleteSchedule)
        QtCore.QMetaObject.connectSlotsByName(DeleteSchedule)

    def retranslateUi(self, DeleteSchedule):
        _translate = QCoreApplication.translate
        DeleteSchedule.setWindowTitle(_translate("DeleteSchedule", "Удаление"))
        self.delete_btn.setText(_translate("DeleteSchedule", "Удалить"))


class DeleteSchedule(QtWidgets.QDialog, Ui_DeleteSchedule):
    def __init__(self, data, parent=None):
        sub_name = []
        tea_name = []
        gr_name = []
        cour_name = []
        ye_name = []
        nl_name = []
        nd_name = []
        tw_name = []
        org_name = []
        ss_name = []
        sub_name.append(data[0])
        tea_name.append(data[1])
        gr_name.append(data[2])
        cour_name.append(data[3])
        ye_name.append(data[4])
        nl_name.append(data[5])
        nd_name.append(data[6])
        tw_name.append(data[7])
        org_name.append(data[8])
        ss_name.append(data[9])
        id_data = []
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        cursor.execute(sel_sub, sub_name)
        id_data.append(cursor.fetchone()[0])
        sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
        cursor.execute(sel_tea, tea_name)
        id_data.append(cursor.fetchone()[0])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_per = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        cursor.execute(sel_per, nl_name)
        id_data.append(cursor.fetchone()[0])
        sel_tt = 'SELECT `id` FROM `name_day` WHERE `name` = %s'
        cursor.execute(sel_tt, nd_name)
        id_data.append(cursor.fetchone()[0])
        sel_bl = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        cursor.execute(sel_bl, tw_name)
        id_data.append(cursor.fetchone()[0])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        sel_ss = 'SELECT `id` FROM `sprav_schedule` WHERE `name` = %s'
        cursor.execute(sel_ss, ss_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[10]))
        main_id = []
        sel_main_id = 'SELECT `id` FROM `schedule` WHERE `subjects_id` = %s AND `teachers_id` = %s AND `groups_id` = %s AND ' \
                      '`courses_id` = %s AND `year_enter_id` = %s AND `num_lessons_id` = %s AND `name_day_id` = %s AND ' \
                      '`type_week_id` = %s AND `organization_id` = %s AND `sprav_schedule_id` = %s AND `num_group` = %s'
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(DeleteSchedule, self).__init__(parent)
        self.setupUi(self)
        self.delete_btn.clicked.connect(self.delete_schedule)
        sch_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch = 'SELECT * FROM `schedule` WHERE `id` = %s'
            cursor.execute(sel_sch, id_)
            sch_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_id[0]) + 1)
        self.table.setRowCount(len(sch_id))
        self.table.setHorizontalHeaderLabels(
            ["Статус", "Номер", "Дисциплина", "Преподаватель", "Группа", "Курсы", "Год поступления", "Номер пары",
             "День недели", "Тип недели", "Организация", "Номер группы", "Основное расписание"])
        self.table.horizontalHeaderItem(1).setToolTip("Column 1")
        self.table.horizontalHeaderItem(2).setToolTip("Column 2")
        self.table.horizontalHeaderItem(3).setToolTip("Column 3")
        self.table.horizontalHeaderItem(4).setToolTip("Column 4")
        self.table.horizontalHeaderItem(5).setToolTip("Column 5")
        self.table.horizontalHeaderItem(6).setToolTip("Column 6")
        self.table.horizontalHeaderItem(7).setToolTip("Column 7")
        self.table.horizontalHeaderItem(8).setToolTip("Column 8")
        self.table.horizontalHeaderItem(9).setToolTip("Column 9")
        self.table.horizontalHeaderItem(10).setToolTip("Column 10")
        self.table.horizontalHeaderItem(11).setToolTip("Column 11")
        self.table.horizontalHeaderItem(12).setToolTip("Column 12")
        self.table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(5).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(6).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(7).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(8).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(9).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(10).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(11).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(12).setTextAlignment(Qt.AlignHCenter)
        for i in range(0, len(sch_id)):
            s_id = []
            s_id.append(sch_id[i][0])
            s_id.append(sch_id[i][1])
            s_id.append(sch_id[i][2])
            s_id.append(sch_id[i][3])
            s_id.append(sch_id[i][4])
            s_id.append(sch_id[i][5])
            s_id.append(sch_id[i][6])
            s_id.append(sch_id[i][7])
            s_id.append(sch_id[i][8])
            s_id.append(sch_id[i][9])
            s_id.append(sch_id[i][10])
            s_id.append(sch_id[i][11])
            sel_sch_name = 'SELECT ' \
                           '(SELECT `id` FROM `schedule` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `teachers` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `num_lessons` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `name_day` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `type_week` WHERE `id` = %s), ' \
                           '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                           '(SELECT `num_group` FROM `schedule` WHERE `num_group` = %s LIMIT 1), ' \
                           '(SELECT `name` FROM `sprav_schedule` WHERE `id` = %s) ' \
                           'FROM `schedule`'
            cursor.execute(sel_sch_name, s_id)
            sch = cursor.fetchone()
            for j in range(0, len(sch)):
                item = QTableWidgetItem(str(sch[j]))
                item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(i, j + 1, item)
            widget = QWidget()
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            layoutH = QHBoxLayout(widget)
            layoutH.addWidget(checkbox)
            layoutH.setAlignment(Qt.AlignCenter)
            layoutH.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(i, 0, widget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.delete_btn)

    def delete_schedule(self):
        checked_list = []
        for i in range(self.table.rowCount()):
            if self.table.cellWidget(i, 0).findChild(type(QCheckBox())).isChecked():
                checked_list.append(self.table.item(i, 1).text())
        for i in range(0, len(checked_list)):
            id_ = []
            id_.append(checked_list[i])
            del_vopr = 'DELETE FROM `schedule` WHERE `id` = %s'
            cursor.execute(del_vopr, id_)
            conn.commit()


class Ui_DeleteScheduleChanges(QtWidgets.QWidget):
    def setupUi(self, DeleteScheduleChanges):
        DeleteScheduleChanges.setObjectName("DeleteScheduleChanges")
        DeleteScheduleChanges.resize(496, 265)
        self.delete_btn = QtWidgets.QPushButton(DeleteScheduleChanges)
        self.delete_btn.setGeometry(QRect(230, 200, 51, 23))
        self.delete_btn.setObjectName("DeleteScheduleChanges")
        self.retranslateUi(DeleteScheduleChanges)
        QtCore.QMetaObject.connectSlotsByName(DeleteScheduleChanges)

    def retranslateUi(self, DeleteScheduleChanges):
        _translate = QCoreApplication.translate
        DeleteScheduleChanges.setWindowTitle(_translate("DeleteScheduleChanges", "Удаление"))
        self.delete_btn.setText(_translate("DeleteScheduleChanges", "Удалить"))


class DeleteScheduleChanges(QtWidgets.QDialog, Ui_DeleteScheduleChanges):
    def __init__(self, data, parent=None):
        gr_name = []
        cour_name = []
        ye_name = []
        nl_name = []
        sub_name = []
        tea_name = []
        org_name = []
        gr_name.append(data[0])
        cour_name.append(data[1])
        ye_name.append(data[2])
        nl_name.append(data[3])
        sub_name.append(data[4])
        tea_name.append(data[5])
        org_name.append(data[7])
        id_data = []
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        sel_nl = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        cursor.execute(sel_nl, nl_name)
        id_data.append(cursor.fetchone()[0])
        if sub_name[0] == '<<Не определено>>' and tea_name[0] == '<<Не определено>>':
            id_data.append(None)
            id_data.append(None)
            sel_main_id = 'SELECT `id` FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                          '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` is %s AND `teachers_id` is %s AND ' \
                          '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
        else:
            sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
            cursor.execute(sel_sub, sub_name)
            id_data.append(cursor.fetchone()[0])
            sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
            cursor.execute(sel_tea, tea_name)
            id_data.append(cursor.fetchone()[0])
            sel_main_id = 'SELECT `id` FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                          '`year_enter_id` = %s AND `num_lessons_id` = %s AND `subjects_id` = %s AND `teachers_id` = %s AND ' \
                          '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
        id_data.append(data[6])
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[8]))
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(DeleteScheduleChanges, self).__init__(parent)
        self.setupUi(self)
        self.delete_btn.clicked.connect(self.delete_schedule_changes)
        sch_c_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch_c = 'SELECT * FROM `schedule_changes` WHERE `id` = %s'
            cursor.execute(sel_sch_c, id_)
            sch_c_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_c_id[0]) + 1)
        self.table.setRowCount(len(sch_c_id))
        self.table.setHorizontalHeaderLabels(
            ["Статус", "Номер", "Группа", "Курсы", "Год поступления", "Номер пары", "Дисциплина", "Преподаватель",  "Дата",
             "Организация", "Номер группы"])
        self.table.horizontalHeaderItem(1).setToolTip("Column 1")
        self.table.horizontalHeaderItem(2).setToolTip("Column 2")
        self.table.horizontalHeaderItem(3).setToolTip("Column 3")
        self.table.horizontalHeaderItem(4).setToolTip("Column 4")
        self.table.horizontalHeaderItem(5).setToolTip("Column 5")
        self.table.horizontalHeaderItem(6).setToolTip("Column 6")
        self.table.horizontalHeaderItem(7).setToolTip("Column 7")
        self.table.horizontalHeaderItem(8).setToolTip("Column 8")
        self.table.horizontalHeaderItem(9).setToolTip("Column 9")
        self.table.horizontalHeaderItem(10).setToolTip("Column 10")
        self.table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(5).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(6).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(7).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(8).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(9).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(10).setTextAlignment(Qt.AlignHCenter)
        for i in range(0, len(sch_c_id)):
            sc_id = []
            sc_id.append(sch_c_id[i][0])
            sc_id.append(sch_c_id[i][1])
            sc_id.append(sch_c_id[i][2])
            sc_id.append(sch_c_id[i][3])
            sc_id.append(sch_c_id[i][4])
            sc_id.append(sch_c_id[i][5])
            sc_id.append(sch_c_id[i][6])
            sc_id.append(sch_c_id[i][7])
            sc_id.append(sch_c_id[i][8])
            sc_id.append(sch_c_id[i][9])
            sel_sch_c_name = 'SELECT ' \
                             '(SELECT `id` FROM `schedule_changes` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `num_lessons` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                             '(SELECT `name` FROM `teachers` WHERE `id` = %s), ' \
                             '(SELECT `date_changes` FROM `schedule_changes` WHERE `date_changes` = %s LIMIT 1), ' \
                             '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                             '(SELECT `num_group` FROM `schedule_changes` WHERE `num_group` = %s LIMIT 1) ' \
                             'FROM `schedule_changes`'
            cursor.execute(sel_sch_c_name, sc_id)
            sch = cursor.fetchone()
            for j in range(0, len(sch)):
                item = QTableWidgetItem(str(sch[j]))
                item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(i, j + 1, item)
            widget = QWidget()
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            layoutH = QHBoxLayout(widget)
            layoutH.addWidget(checkbox)
            layoutH.setAlignment(Qt.AlignCenter)
            layoutH.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(i, 0, widget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.delete_btn)

    def delete_schedule_changes(self):
        checked_list = []
        for i in range(self.table.rowCount()):
            if self.table.cellWidget(i, 0).findChild(type(QCheckBox())).isChecked():
                checked_list.append(self.table.item(i, 1).text())
        for i in range(0, len(checked_list)):
            id_ = []
            id_.append(checked_list[i])
            del_vopr = 'DELETE FROM `schedule_changes` WHERE `id` = %s'
            cursor.execute(del_vopr, id_)
            conn.commit()


class Ui_DeleteThemeHours(QtWidgets.QWidget):
    def setupUi(self, DeleteThemeHours):
        DeleteThemeHours.setObjectName("DeleteThemeHours")
        DeleteThemeHours.resize(496, 265)
        self.delete_btn = QtWidgets.QPushButton(DeleteThemeHours)
        self.delete_btn.setGeometry(QRect(230, 200, 51, 23))
        self.delete_btn.setObjectName("DeleteThemeHours")
        self.retranslateUi(DeleteThemeHours)
        QtCore.QMetaObject.connectSlotsByName(DeleteThemeHours)

    def retranslateUi(self, DeleteThemeHours):
        _translate = QCoreApplication.translate
        DeleteThemeHours.setWindowTitle(_translate("DeleteThemeHours", "Удаление"))
        self.delete_btn.setText(_translate("DeleteThemeHours", "Удалить"))


class DeleteThemeHours(QtWidgets.QDialog, Ui_DeleteThemeHours):
    def __init__(self, data, parent=None):
        sub_name = []
        gr_name = []
        cour_name = []
        ye_name = []
        org_name = []
        sub_name.append(data[0])
        gr_name.append(data[1])
        cour_name.append(data[2])
        ye_name.append(data[3])
        org_name.append(data[4])
        id_data = []
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        cursor.execute(sel_sub, sub_name)
        id_data.append(cursor.fetchone()[0])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_gr, gr_name)
        id_data.append(cursor.fetchone()[0])
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_cour, cour_name)
        id_data.append(cursor.fetchone()[0])
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_ye, ye_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[5]))
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_org, org_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(int(data[6]))
        sel_main_id = 'SELECT `id` FROM `lessons_plan` WHERE `subjects_id` = %s AND `groups_id` = %s AND `courses_id` ' \
                      '= %s AND `year_enter_id` = %s AND `number` = %s AND `organization_id` = %s AND `term` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(DeleteThemeHours, self).__init__(parent)
        self.setupUi(self)
        self.delete_btn.clicked.connect(self.delete_theme_hours)
        sch_c_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch_c = 'SELECT * FROM `lessons_plan` WHERE `id` = %s'
            cursor.execute(sel_sch_c, id_)
            sch_c_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_c_id[0]) + 1)
        self.table.setRowCount(len(sch_c_id))
        self.table.setHorizontalHeaderLabels(
            ["Статус", "Номер", "Дисциплина", "Тема", "Группа", "Курсы", "Год поступления", "Количество часов",
             "Организация", "Семестр"])
        self.table.horizontalHeaderItem(1).setToolTip("Column 1")
        self.table.horizontalHeaderItem(2).setToolTip("Column 2")
        self.table.horizontalHeaderItem(3).setToolTip("Column 3")
        self.table.horizontalHeaderItem(4).setToolTip("Column 4")
        self.table.horizontalHeaderItem(5).setToolTip("Column 5")
        self.table.horizontalHeaderItem(6).setToolTip("Column 6")
        self.table.horizontalHeaderItem(7).setToolTip("Column 7")
        self.table.horizontalHeaderItem(8).setToolTip("Column 8")
        self.table.horizontalHeaderItem(9).setToolTip("Column 9")
        self.table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(4).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(5).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(6).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(7).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(8).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(9).setTextAlignment(Qt.AlignHCenter)
        for i in range(0, len(sch_c_id)):
            sel_th_name = 'SELECT ' \
                          '(SELECT `id` FROM `lessons_plan` WHERE `id` = %s), ' \
                          '(SELECT `name` FROM `subjects` WHERE `id` = %s), ' \
                          '(SELECT `theme` FROM `lessons_plan` WHERE `theme` = %s LIMIT 1), ' \
                          '(SELECT `name` FROM `groups` WHERE `id` = %s), ' \
                          '(SELECT `name` FROM `courses` WHERE `id` = %s), ' \
                          '(SELECT `name` FROM `year_enter` WHERE `id` = %s), ' \
                          '(SELECT `number` FROM `lessons_plan` WHERE `number` = %s LIMIT 1), ' \
                          '(SELECT `name` FROM `organization` WHERE `id` = %s), ' \
                          '(SELECT `term` FROM `lessons_plan` WHERE `term` = %s LIMIT 1) ' \
                          'FROM `lessons_plan`'
            cursor.execute(sel_th_name, sch_c_id[i])
            sch = cursor.fetchone()
            for j in range(0, len(sch)):
                item = QTableWidgetItem(str(sch[j]))
                item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(i, j + 1, item)
            widget = QWidget()
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            layoutH = QHBoxLayout(widget)
            layoutH.addWidget(checkbox)
            layoutH.setAlignment(Qt.AlignCenter)
            layoutH.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(i, 0, widget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.delete_btn)

    def delete_theme_hours(self):
        checked_list = []
        for i in range(self.table.rowCount()):
            if self.table.cellWidget(i, 0).findChild(type(QCheckBox())).isChecked():
                checked_list.append(self.table.item(i, 1).text())
        for i in range(0, len(checked_list)):
            id_ = []
            id_.append(checked_list[i])
            del_vopr = 'DELETE FROM `lessons_plan` WHERE `id` = %s'
            cursor.execute(del_vopr, id_)
            conn.commit()


class Ui_DeleteWeekDate(QtWidgets.QWidget):
    def setupUi(self, DeleteWeekDate):
        DeleteWeekDate.setObjectName("DeleteWeekDate")
        DeleteWeekDate.resize(496, 265)
        self.delete_btn = QtWidgets.QPushButton(DeleteWeekDate)
        self.delete_btn.setGeometry(QRect(230, 200, 51, 23))
        self.delete_btn.setObjectName("DeleteWeekDate")
        self.retranslateUi(DeleteWeekDate)
        QtCore.QMetaObject.connectSlotsByName(DeleteWeekDate)

    def retranslateUi(self, DeleteWeekDate):
        _translate = QCoreApplication.translate
        DeleteWeekDate.setWindowTitle(_translate("DeleteWeekDate", "Удаление"))
        self.delete_btn.setText(_translate("DeleteWeekDate", "Удалить"))


class DeleteWeekDate(QtWidgets.QDialog, Ui_DeleteWeekDate):
    def __init__(self, data, parent=None):
        tt_name = []
        tt_name.append(data[0])
        id_data = []
        sel_sub = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        cursor.execute(sel_sub, tt_name)
        id_data.append(cursor.fetchone()[0])
        id_data.append(data[1])
        sel_main_id = 'SELECT `id` FROM `date_type_week` WHERE `type_week_id` = %s AND `date_week` = %s'
        main_id = []
        cursor.execute(sel_main_id, id_data)
        id_ = cursor.fetchall()
        for i in range(0, len(id_)):
            main_id.append(id_[i][0])
        super(DeleteWeekDate, self).__init__(parent)
        self.setupUi(self)
        self.delete_btn.clicked.connect(self.delete_week_date)
        sch_c_id = []
        for i in range(0, len(main_id)):
            id_ = []
            id_.append(main_id[i])
            sel_sch_c = 'SELECT * FROM `date_type_week` WHERE `id` = %s'
            cursor.execute(sel_sch_c, id_)
            sch_c_id.append(cursor.fetchall()[0])
        self.table = QTableWidget(self)
        self.table.setColumnCount(len(sch_c_id[0]) + 1)
        self.table.setRowCount(len(sch_c_id))
        self.table.setHorizontalHeaderLabels(
            ["Статус", "Номер", "Тип недели", "Дата"])
        self.table.horizontalHeaderItem(1).setToolTip("Column 1")
        self.table.horizontalHeaderItem(2).setToolTip("Column 2")
        self.table.horizontalHeaderItem(3).setToolTip("Column 3")
        self.table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        self.table.horizontalHeaderItem(3).setTextAlignment(Qt.AlignHCenter)
        for i in range(0, len(sch_c_id)):
            sel_wd_name = 'SELECT ' \
                          '(SELECT `id` FROM `date_type_week` WHERE `id` = %s), ' \
                          '(SELECT `name` FROM `type_week` WHERE `id` = %s), ' \
                          '(SELECT `date_week` FROM `date_type_week` WHERE `date_week` = %s LIMIT 1) ' \
                          'FROM `date_type_week`'
            cursor.execute(sel_wd_name, sch_c_id[i])
            sch = cursor.fetchone()
            for j in range(0, len(sch)):
                item = QTableWidgetItem(str(sch[j]))
                item.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(i, j + 1, item)
            widget = QWidget()
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            layoutH = QHBoxLayout(widget)
            layoutH.addWidget(checkbox)
            layoutH.setAlignment(Qt.AlignCenter)
            layoutH.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(i, 0, widget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.table)
        vlayout.addWidget(self.delete_btn)

    def delete_week_date(self):
        checked_list = []
        for i in range(self.table.rowCount()):
            if self.table.cellWidget(i, 0).findChild(type(QCheckBox())).isChecked():
                checked_list.append(self.table.item(i, 1).text())
        for i in range(0, len(checked_list)):
            id_ = []
            id_.append(checked_list[i])
            del_vopr = 'DELETE FROM `date_type_week` WHERE `id` = %s'
            cursor.execute(del_vopr, id_)
            conn.commit()


class Ui_signUp(QtWidgets.QWidget):
    def setupUi(self, Dialog):
        super(Ui_signUp, self).__init__()
        Dialog.setObjectName("Dialog")
        Dialog.resize(570, 375)
        self.uname_lineEdit = QtWidgets.QLineEdit()
        self.password_lineEdit = QtWidgets.QLineEdit()
        self.connect = QtWidgets.QPushButton("connect")
        self.signup_btn = QPushButton(Dialog)
        self.signup_btn.setObjectName("signup_btn")
        self.signup_btn.clicked.connect(Dialog.insertData)
        self.verticalLayoutWidget = QtWidgets.QWidget(Dialog)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(10, 10, 441, 321))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.combo = QtWidgets.QComboBox(self.verticalLayoutWidget)
        self.combo.setObjectName("combo")
        self.combo1 = QtWidgets.QComboBox(self.verticalLayoutWidget)
        self.combo1.setObjectName("combo1")
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label.setObjectName("label")
        self.label1 = QtWidgets.QLabel(Dialog)
        self.label1.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label1.setObjectName("label1")
        self.label2 = QtWidgets.QLabel(Dialog)
        self.label2.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label2.setObjectName("label2")
        self.label3 = QtWidgets.QLabel(Dialog)
        self.label3.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label3.setObjectName("label3")
        cursor.execute('SELECT `name` FROM `roles`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo.addItem(check_sel[i][0])
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo1.addItem(check_sel[i][0])
        vl = QVBoxLayout()
        vl.addWidget(self.label)
        vl.addWidget(self.uname_lineEdit)
        vl.addWidget(self.label1)
        vl.addWidget(self.password_lineEdit)
        vl.addWidget(self.label2)
        vl.addWidget(self.combo)
        vl.addWidget(self.label3)
        vl.addWidget(self.combo1)
        vl.addWidget(self.signup_btn)
        self.setLayout(vl)
        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Регистрация"))
        self.signup_btn.setText(_translate("Dialog", "Зарегистрировать"))
        self.label.setText(_translate("Dialog", "Логин:"))
        self.label1.setText(_translate("Dialog", "Пароль:"))
        self.label2.setText(_translate("Dialog", "Роль:"))
        self.label3.setText(_translate("Dialog", "Преподаватель:"))


class Dialog(QDialog, Ui_signUp):
    def __init__(self, parent=None):
        super(Dialog, self).__init__(parent)
        self.setupUi(self)
        self.parent = parent
        self.setModal(True)
        self.signup_btn.clicked.connect(self.insertData)

    @pyqtSlot()
    def insertData(self):

        username = self.uname_lineEdit.text()
        password = self.password_lineEdit.text()
        role = self.combo.currentText()
        if role == 'Преподаватель':
            teachers = self.combo1.currentText()
        else:
            teachers = None
        if len(password) < 8:
            msg = QMessageBox.information(self, 'Внимание!', 'Пароль должен иметь минимум 8 символов.')
            return
        if (not username) or (not password):
            msg = QMessageBox.information(self, 'Внимание!', 'Вы не заполнили все поля.')
            return
        sel_user = 'SELECT * FROM `users` WHERE `login` = %s'
        user = []
        user.append(username)
        cursor.execute(sel_user, user)
        if cursor.fetchone() != None:
            msg = QMessageBox.information(self, 'Внимание!', 'Пользоватеть с таким именем уже зарегистрирован.')
        else:
            sel_role_id = 'SELECT `id` FROM `roles` WHERE `name` = %s'
            role_name = []
            role_name.append(role)
            cursor.execute(sel_role_id, role_name)
            role_id = cursor.fetchone()[0]
            if teachers != None:
                sel_teacher_id = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
                teacher_name = []
                teacher_name.append(teachers)
                cursor.execute(sel_teacher_id, teacher_name)
                teacher_id = cursor.fetchone()[0]
            else:
                teacher_id = None
            in_user = "INSERT INTO `users` (`login`, `password`, `roles_id`, `teachers_id`) VALUES (%s, %s, %s, %s)"
            user.append(password)
            user.append(role_id)
            user.append(teacher_id)
            cursor.execute(in_user, user)
            conn.commit()

    def closeEvent(self, event):
        self.parent.show()


class Ui_Login(QtWidgets.QWidget):
    def setupUi(self, Login):
        Login.setObjectName("Login")
        Login.resize(496, 265)
        self.uname_lineEdit = QLineEdit(Login)
        self.uname_lineEdit.setGeometry(QRect(230, 110, 113, 20))
        self.uname_lineEdit.setObjectName("uname_lineEdit")
        self.pass_lineEdit = QLineEdit(Login)
        self.pass_lineEdit.setGeometry(QRect(230, 150, 113, 20))
        self.pass_lineEdit.setObjectName("pass_lineEdit")
        self.pass_lineEdit.setEchoMode(QtWidgets.QLineEdit.EchoMode.Password)
        self.login_btn = QPushButton(Login)
        self.login_btn.setGeometry(QRect(230, 200, 51, 23))
        self.login_btn.setObjectName("login_btn")
        self.signup_btn = QPushButton(Login)
        self.signup_btn.setGeometry(QRect(290, 200, 51, 23))
        self.signup_btn.setObjectName("signup_btn")
        self.label = QLabel(Login)
        self.label.setGeometry(QRect(190, 10, 211, 51))
        self.label.setObjectName("label")
        self.label1 = QLabel(Login)
        self.label1.setGeometry(QRect(190, 10, 211, 51))
        self.label1.setObjectName("label1")
        hl = QHBoxLayout()
        hl.addWidget(self.label)
        hl.addWidget(self.uname_lineEdit)
        hl1 = QHBoxLayout()
        hl1.addWidget(self.label1)
        hl1.addWidget(self.pass_lineEdit)
        hl2 = QHBoxLayout()
        hl2.addWidget(self.login_btn)
        hl2.addWidget(self.signup_btn)
        vl = QVBoxLayout()
        vl.addLayout(hl)
        vl.addLayout(hl1)
        vl.addLayout(hl2)
        self.setLayout(vl)
        self.retranslateUi(Login)
        QMetaObject.connectSlotsByName(Login)

    def retranslateUi(self, Login):
        _translate = QCoreApplication.translate
        Login.setWindowTitle(_translate("Login", "Вход в систему"))
        self.login_btn.setText(_translate("Login", "Войти"))
        self.signup_btn.setText(_translate("Login", "Зарегистрировать"))
        self.label.setText(_translate("Login", "Логин"))
        self.label1.setText(_translate("Login", "Пароль"))


class Login(QtWidgets.QDialog, Ui_Login):
    # основная логика окна
    def __init__(self, parent=None):
        super(Login, self).__init__(parent)
        self.setupUi(self)
        self.login_btn.clicked.connect(self.loginCheck)
        self.signup_btn.clicked.connect(self.signUpCheck)

    # показать сообщения
    def showMessageBox(self, title, message):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Warning)
        msgBox.setWindowTitle(title)
        msgBox.setText(message)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    # открыть класс главного окна
    def mainWindowShow(self, login):
        self.mainWindow = MainWindow(login)
        self.show()

    # проверка авторизации
    def loginCheck(self):
        username = self.uname_lineEdit.text()
        password = self.pass_lineEdit.text()
        if (not username) or (not password):
            msg = QMessageBox.information(self, 'Внимание!', 'Вы не заполнили все поля.')
            return
        sel_user = 'SELECT * FROM `users` WHERE `login` = %s AND `password` = %s'
        user = []
        user.append(username)
        user.append(password)
        cursor.execute(sel_user, user)
        if cursor.fetchone() != None:
            self.mainWindowShow(username)
            self.hide()
        else:
            self.showMessageBox('Внимание!', 'Неправильное имя пользователя или пароль.')

    # открыть класс с созданием нового пользователя
    def signUpShow(self):
        self.signUpWindow = Dialog(self)
        self.signUpWindow.show()

    # проверка админ прав
    def is_admin(self):
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    # проверка регистрации
    def signUpCheck(self):
        if self.is_admin():
            self.signUpShow()
        else:
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__, None, 1)


class SimpleDateValidator(QtGui.QValidator):
    def validate(self, text, pos):
        if not text:
            return self.Acceptable, text, pos
        fmt = self.parent().format()
        _sep = set(fmt.replace('d', '').replace('M', '').replace('y', ''))

        for l in text:
            # убедитесь, что набранный текст представляет собой цифру или разделитель
            if not l.isdigit() and l not in _sep:
                return self.Invalid, text, pos
        years = fmt.count('y')
        if len(text) <= years and text.isdigit():
            return self.Acceptable, text, pos
        if QtCore.QDate.fromString(text, fmt).isValid():
            return self.Acceptable, text, pos
        return self.Intermediate, text, pos


class Ui_MainWindow(QtWidgets.QWidget):
    # объявление всех кнопок и надписей
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        self.widget = QWidget()
        self.textBox = QPlainTextEdit(self.widget)
        self.textBox.move(250, 120)
        self.u_textBox = QPlainTextEdit(self.widget)
        self.u_textBox.move(250, 120)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.open_pars_date_week = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_date_week.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_date_week.setObjectName("open_pars_date_week")
        self.open_pars_date_week.findChild(QPushButton, 'open_pars_date_week')
        self.open_pars_s = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_s.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_s.setObjectName("open_pars_s")
        self.open_pars_s.findChild(QPushButton, 'open_pars_s')
        self.open_pars_sc = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_sc.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_sc.setObjectName("open_pars_sc")
        self.open_pars_sc.findChild(QPushButton, 'open_pars_sc')
        self.open_pars_th = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_th.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_th.setObjectName("open_pars_th")
        self.open_pars_th.findChild(QPushButton, 'open_pars_th')
        self.open_pars_sub = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_sub.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_sub.setObjectName("open_pars_sub")
        self.open_pars_sub.findChild(QPushButton, 'open_pars_sub')
        self.open_pars_tea = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_tea.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_tea.setObjectName("open_pars_tea")
        self.open_pars_tea.findChild(QPushButton, 'open_pars_tea')
        self.open_pars_stud = QtWidgets.QPushButton(self.centralwidget)
        self.open_pars_stud.setGeometry(QtCore.QRect(250, 130, 89, 25))
        self.open_pars_stud.setObjectName("open_pars_stud")
        self.open_pars_stud.findChild(QPushButton, 'open_pars_stud')
        self.pars_s = QtWidgets.QPushButton(self.centralwidget)
        self.pars_s.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_s.setObjectName("pars_s")
        self.pars_s.findChild(QPushButton, 'pars_s')
        self.pars_sc = QtWidgets.QPushButton(self.centralwidget)
        self.pars_sc.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_sc.setObjectName("pars_sc")
        self.pars_sc.findChild(QPushButton, 'pars_sc')
        self.pars_th = QtWidgets.QPushButton(self.centralwidget)
        self.pars_th.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_th.setObjectName("pars_th")
        self.pars_th.findChild(QPushButton, 'pars_th')
        self.pars_sub = QtWidgets.QPushButton(self.centralwidget)
        self.pars_sub.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_sub.setObjectName("pars_sub")
        self.pars_sub.findChild(QPushButton, 'pars_sub')
        self.pars_tea = QtWidgets.QPushButton(self.centralwidget)
        self.pars_tea.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_tea.setObjectName("pars_tea")
        self.pars_tea.findChild(QPushButton, 'pars_tea')
        self.pars_stud = QtWidgets.QPushButton(self.centralwidget)
        self.pars_stud.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_stud.setObjectName("pars_stud")
        self.pars_stud.findChild(QPushButton, 'pars_stud')
        self.pars_date_week = QtWidgets.QPushButton(self.centralwidget)
        self.pars_date_week.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.pars_date_week.setObjectName("pars_date_week")
        self.pars_date_week.findChild(QPushButton, 'pars_date_week')
        self.label_21 = QtWidgets.QLabel(self.centralwidget)
        self.label_21.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_21.setObjectName("label_21")
        self.label_22 = QtWidgets.QLabel(self.centralwidget)
        self.label_22.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_22.setObjectName("label_22")
        self.label_23 = QtWidgets.QLabel(self.centralwidget)
        self.label_23.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_23.setObjectName("label_23")
        self.label_24 = QtWidgets.QLabel(self.centralwidget)
        self.label_24.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_24.setObjectName("label_24")
        self.label_25 = QtWidgets.QLabel(self.centralwidget)
        self.label_25.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_25.setObjectName("label_25")
        self.label_26 = QtWidgets.QLabel(self.centralwidget)
        self.label_26.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_26.setObjectName("label_26")
        self.label_27 = QtWidgets.QLabel(self.centralwidget)
        self.label_27.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_27.setObjectName("label_27")
        self.label_28 = QtWidgets.QLabel(self.centralwidget)
        self.label_28.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_28.setObjectName("label_28")
        self.label_29 = QtWidgets.QLabel(self.centralwidget)
        self.label_29.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_29.setObjectName("label_29")
        self.label_30 = QtWidgets.QLabel(self.centralwidget)
        self.label_30.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_30.setObjectName("label_30")
        self.label_31 = QtWidgets.QLabel(self.centralwidget)
        self.label_31.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_31.setObjectName("label_31")
        self.label_71 = QtWidgets.QLabel(self.centralwidget)
        self.label_71.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_71.setObjectName("label_71")
        self.label_72 = QtWidgets.QLabel(self.centralwidget)
        self.label_72.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_72.setObjectName("label_72")
        self.label_73 = QtWidgets.QLabel(self.centralwidget)
        self.label_73.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_73.setObjectName("label_73")
        self.label_74 = QtWidgets.QLabel(self.centralwidget)
        self.label_74.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_74.setObjectName("label_74")
        self.label_75 = QtWidgets.QLabel(self.centralwidget)
        self.label_75.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_75.setObjectName("label_75")
        self.label_76 = QtWidgets.QLabel(self.centralwidget)
        self.label_76.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_76.setObjectName("label_76")
        self.label_77 = QtWidgets.QLabel(self.centralwidget)
        self.label_77.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_77.setObjectName("label_77")
        self.label_78 = QtWidgets.QLabel(self.centralwidget)
        self.label_78.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_78.setObjectName("label_78")
        self.label_79 = QtWidgets.QLabel(self.centralwidget)
        self.label_79.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_79.setObjectName("label_79")
        self.label_80 = QtWidgets.QLabel(self.centralwidget)
        self.label_80.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_80.setObjectName("label_80")
        self.label_81 = QtWidgets.QLabel(self.centralwidget)
        self.label_81.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_81.setObjectName("label_81")
        self.label_121 = QtWidgets.QLabel(self.centralwidget)
        self.label_121.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_121.setObjectName("label_121")
        self.label_122 = QtWidgets.QLabel(self.centralwidget)
        self.label_122.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_122.setObjectName("label_122")
        self.label_123 = QtWidgets.QLabel(self.centralwidget)
        self.label_123.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_123.setObjectName("label_123")
        self.label_124 = QtWidgets.QLabel(self.centralwidget)
        self.label_124.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_124.setObjectName("label_124")
        self.label_125 = QtWidgets.QLabel(self.centralwidget)
        self.label_125.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_125.setObjectName("label_125")
        self.label_126 = QtWidgets.QLabel(self.centralwidget)
        self.label_126.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_126.setObjectName("label_126")
        self.label_127 = QtWidgets.QLabel(self.centralwidget)
        self.label_127.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_127.setObjectName("label_127")
        self.label_128 = QtWidgets.QLabel(self.centralwidget)
        self.label_128.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_128.setObjectName("label_128")
        self.label_129 = QtWidgets.QLabel(self.centralwidget)
        self.label_129.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_129.setObjectName("label_129")
        self.label_130 = QtWidgets.QLabel(self.centralwidget)
        self.label_130.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_130.setObjectName("label_130")
        self.label_131 = QtWidgets.QLabel(self.centralwidget)
        self.label_131.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_131.setObjectName("label_131")
        self.label_40 = QtWidgets.QLabel(self.centralwidget)
        self.label_40.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_40.setObjectName("label_40")
        self.label_41 = QtWidgets.QLabel(self.centralwidget)
        self.label_41.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_41.setObjectName("label_41")
        self.label_42 = QtWidgets.QLabel(self.centralwidget)
        self.label_42.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_42.setObjectName("label_42")
        self.label_43 = QtWidgets.QLabel(self.centralwidget)
        self.label_43.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_43.setObjectName("label_43")
        self.label_44 = QtWidgets.QLabel(self.centralwidget)
        self.label_44.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_44.setObjectName("label_44")
        self.label_45 = QtWidgets.QLabel(self.centralwidget)
        self.label_45.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_45.setObjectName("label_45")
        self.label_46 = QtWidgets.QLabel(self.centralwidget)
        self.label_46.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_46.setObjectName("label_46")
        self.label_47 = QtWidgets.QLabel(self.centralwidget)
        self.label_47.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_47.setObjectName("label_47")
        self.label_48 = QtWidgets.QLabel(self.centralwidget)
        self.label_48.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_48.setObjectName("label_48")
        self.label_140 = QtWidgets.QLabel(self.centralwidget)
        self.label_140.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_140.setObjectName("label_140")
        self.label_141 = QtWidgets.QLabel(self.centralwidget)
        self.label_141.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_141.setObjectName("label_141")
        self.label_142 = QtWidgets.QLabel(self.centralwidget)
        self.label_142.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_142.setObjectName("label_142")
        self.label_143 = QtWidgets.QLabel(self.centralwidget)
        self.label_143.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_143.setObjectName("label_143")
        self.label_144 = QtWidgets.QLabel(self.centralwidget)
        self.label_144.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_144.setObjectName("label_144")
        self.label_145 = QtWidgets.QLabel(self.centralwidget)
        self.label_145.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_145.setObjectName("label_145")
        self.label_146 = QtWidgets.QLabel(self.centralwidget)
        self.label_146.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_146.setObjectName("label_146")
        self.label_147 = QtWidgets.QLabel(self.centralwidget)
        self.label_147.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_147.setObjectName("label_147")
        self.label_148 = QtWidgets.QLabel(self.centralwidget)
        self.label_148.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_148.setObjectName("label_148")
        self.label_90 = QtWidgets.QLabel(self.centralwidget)
        self.label_90.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_90.setObjectName("label_90")
        self.label_91 = QtWidgets.QLabel(self.centralwidget)
        self.label_91.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_91.setObjectName("label_91")
        self.label_92 = QtWidgets.QLabel(self.centralwidget)
        self.label_92.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_92.setObjectName("label_92")
        self.label_93 = QtWidgets.QLabel(self.centralwidget)
        self.label_93.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_93.setObjectName("label_93")
        self.label_94 = QtWidgets.QLabel(self.centralwidget)
        self.label_94.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_94.setObjectName("label_94")
        self.label_95 = QtWidgets.QLabel(self.centralwidget)
        self.label_95.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_95.setObjectName("label_95")
        self.label_96 = QtWidgets.QLabel(self.centralwidget)
        self.label_96.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_96.setObjectName("label_96")
        self.label_97 = QtWidgets.QLabel(self.centralwidget)
        self.label_97.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_97.setObjectName("label_97")
        self.label_98 = QtWidgets.QLabel(self.centralwidget)
        self.label_98.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_98.setObjectName("label_98")
        self.label_50 = QtWidgets.QLabel(self.centralwidget)
        self.label_50.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_50.setObjectName("label_50")
        self.label_51 = QtWidgets.QLabel(self.centralwidget)
        self.label_51.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_51.setObjectName("label_51")
        self.label_52 = QtWidgets.QLabel(self.centralwidget)
        self.label_52.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_52.setObjectName("label_52")
        self.label_53 = QtWidgets.QLabel(self.centralwidget)
        self.label_53.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_53.setObjectName("label_53")
        self.label_54 = QtWidgets.QLabel(self.centralwidget)
        self.label_54.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_54.setObjectName("label_54")
        self.label_55 = QtWidgets.QLabel(self.centralwidget)
        self.label_55.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_55.setObjectName("label_55")
        self.label_56 = QtWidgets.QLabel(self.centralwidget)
        self.label_56.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_56.setObjectName("label_56")
        self.label_100 = QtWidgets.QLabel(self.centralwidget)
        self.label_100.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_100.setObjectName("label_100")
        self.label_101 = QtWidgets.QLabel(self.centralwidget)
        self.label_101.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_101.setObjectName("label_101")
        self.label_102 = QtWidgets.QLabel(self.centralwidget)
        self.label_102.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_102.setObjectName("label_102")
        self.label_103 = QtWidgets.QLabel(self.centralwidget)
        self.label_103.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_103.setObjectName("label_103")
        self.label_104 = QtWidgets.QLabel(self.centralwidget)
        self.label_104.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_104.setObjectName("label_104")
        self.label_105 = QtWidgets.QLabel(self.centralwidget)
        self.label_105.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_105.setObjectName("label_105")
        self.label_106 = QtWidgets.QLabel(self.centralwidget)
        self.label_106.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_106.setObjectName("label_106")
        self.label_150 = QtWidgets.QLabel(self.centralwidget)
        self.label_150.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_150.setObjectName("label_150")
        self.label_151 = QtWidgets.QLabel(self.centralwidget)
        self.label_151.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_151.setObjectName("label_151")
        self.label_152 = QtWidgets.QLabel(self.centralwidget)
        self.label_152.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_152.setObjectName("label_152")
        self.label_153 = QtWidgets.QLabel(self.centralwidget)
        self.label_153.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_153.setObjectName("label_153")
        self.label_154 = QtWidgets.QLabel(self.centralwidget)
        self.label_154.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_154.setObjectName("label_154")
        self.label_155 = QtWidgets.QLabel(self.centralwidget)
        self.label_155.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_155.setObjectName("label_155")
        self.label_156 = QtWidgets.QLabel(self.centralwidget)
        self.label_156.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_156.setObjectName("label_156")
        self.label_160 = QtWidgets.QLabel(self.centralwidget)
        self.label_160.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_160.setObjectName("label_160")
        self.label_161 = QtWidgets.QLabel(self.centralwidget)
        self.label_161.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_161.setObjectName("label_161")
        self.label_170 = QtWidgets.QLabel(self.centralwidget)
        self.label_170.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_170.setObjectName("label_170")
        self.label_171 = QtWidgets.QLabel(self.centralwidget)
        self.label_171.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_171.setObjectName("label_171")
        self.label_172 = QtWidgets.QLabel(self.centralwidget)
        self.label_172.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_172.setObjectName("label_172")
        self.label_173 = QtWidgets.QLabel(self.centralwidget)
        self.label_173.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_173.setObjectName("label_173")
        self.label_174 = QtWidgets.QLabel(self.centralwidget)
        self.label_174.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_174.setObjectName("label_174")
        self.label_175 = QtWidgets.QLabel(self.centralwidget)
        self.label_175.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_175.setObjectName("label_175")
        self.label_176 = QtWidgets.QLabel(self.centralwidget)
        self.label_176.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_176.setObjectName("label_176")
        self.label_177 = QtWidgets.QLabel(self.centralwidget)
        self.label_177.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_177.setObjectName("label_177")
        self.label_178 = QtWidgets.QLabel(self.centralwidget)
        self.label_178.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_178.setObjectName("label_178")
        self.label_179 = QtWidgets.QLabel(self.centralwidget)
        self.label_179.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_179.setObjectName("label_179")
        self.label_180 = QtWidgets.QLabel(self.centralwidget)
        self.label_180.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_180.setObjectName("label_180")
        self.label_181 = QtWidgets.QLabel(self.centralwidget)
        self.label_181.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_181.setObjectName("label_181")
        self.label_182 = QtWidgets.QLabel(self.centralwidget)
        self.label_182.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_182.setObjectName("label_182")
        self.label_183 = QtWidgets.QLabel(self.centralwidget)
        self.label_183.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_183.setObjectName("label_183")
        self.label_190 = QtWidgets.QLabel(self.centralwidget)
        self.label_190.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_190.setObjectName("label_190")
        self.label_191 = QtWidgets.QLabel(self.centralwidget)
        self.label_191.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_191.setObjectName("label_191")
        self.label_192 = QtWidgets.QLabel(self.centralwidget)
        self.label_192.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_192.setObjectName("label_192")
        self.label_193 = QtWidgets.QLabel(self.centralwidget)
        self.label_193.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_193.setObjectName("label_193")
        self.label_194 = QtWidgets.QLabel(self.centralwidget)
        self.label_194.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_194.setObjectName("label_194")
        self.label_195 = QtWidgets.QLabel(self.centralwidget)
        self.label_195.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_195.setObjectName("label_195")
        self.label_196 = QtWidgets.QLabel(self.centralwidget)
        self.label_196.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_196.setObjectName("label_196")
        self.label_200 = QtWidgets.QLabel(self.centralwidget)
        self.label_200.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_200.setObjectName("label_200")
        self.label_201 = QtWidgets.QLabel(self.centralwidget)
        self.label_201.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_201.setObjectName("label_201")
        self.label_202 = QtWidgets.QLabel(self.centralwidget)
        self.label_202.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_202.setObjectName("label_202")
        self.label_203 = QtWidgets.QLabel(self.centralwidget)
        self.label_203.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_203.setObjectName("label_203")
        self.label_210 = QtWidgets.QLabel(self.centralwidget)
        self.label_210.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_210.setObjectName("label_210")
        self.label_211 = QtWidgets.QLabel(self.centralwidget)
        self.label_211.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_211.setObjectName("label_211")
        self.label_212 = QtWidgets.QLabel(self.centralwidget)
        self.label_212.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_212.setObjectName("label_212")
        self.label_213 = QtWidgets.QLabel(self.centralwidget)
        self.label_213.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_213.setObjectName("label_213")
        self.label_220 = QtWidgets.QLabel(self.centralwidget)
        self.label_220.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_220.setObjectName("label_220")
        self.label_221 = QtWidgets.QLabel(self.centralwidget)
        self.label_221.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_221.setObjectName("label_221")
        self.label_222 = QtWidgets.QLabel(self.centralwidget)
        self.label_222.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_222.setObjectName("label_222")
        self.label_60 = QtWidgets.QLabel(self.centralwidget)
        self.label_60.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_60.setObjectName("label_60")
        self.label_61 = QtWidgets.QLabel(self.centralwidget)
        self.label_61.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_61.setObjectName("label_61")
        self.label_110 = QtWidgets.QLabel(self.centralwidget)
        self.label_110.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_110.setObjectName("label_110")
        self.label_111 = QtWidgets.QLabel(self.centralwidget)
        self.label_111.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_111.setObjectName("label_111")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_2.setObjectName("label_2")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label.setObjectName("label")
        self.filename = QtWidgets.QLineEdit(self.centralwidget)
        self.filename.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename.setObjectName("filename")
        self.label1 = QtWidgets.QLabel(self.centralwidget)
        self.label1.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label1.setObjectName("label1")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(10, 110, 241, 17))
        self.label_6.setObjectName("label_6")
        self.filename1 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename1.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename1.setObjectName("filename1")
        self.filename_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename_2.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename_2.setObjectName("filename_2")
        self.filename_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename_3.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename_3.setObjectName("filename_3")
        self.filename_4 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename_4.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename_4.setObjectName("filename_4")
        self.filename_5 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename_5.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename_5.setObjectName("filename_5")
        self.filename_6 = QtWidgets.QLineEdit(self.centralwidget)
        self.filename_6.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.filename_6.setObjectName("filename_6")
        self.schedule_line = QtWidgets.QLineEdit(self.centralwidget)
        self.schedule_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.schedule_line.setObjectName("schedule_line")
        self.d_schedule_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_schedule_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_schedule_line.setObjectName("d_schedule_line")
        self.u_schedule_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_schedule_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_schedule_line.setObjectName("u_schedule_line")
        self.s_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.s_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.s_date_line.setObjectName("s_date_line")
        self.sc_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.sc_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.sc_date_line.setObjectName("sc_date_line")
        self.sc_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.u_sc_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_sc_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_sc_date_line.setObjectName("u_sc_date_line")
        self.u_sc_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.d_sc_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_sc_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_sc_date_line.setObjectName("d_sc_date_line")
        self.d_sc_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.wd_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.wd_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.wd_date_line.setObjectName("wd_date_line")
        self.wd_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.u_wd_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_wd_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_wd_date_line.setObjectName("u_wd_date_line")
        self.u_wd_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.d_wd_date_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_wd_date_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_wd_date_line.setObjectName("d_wd_date_line")
        self.d_wd_date_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.sc_line = QtWidgets.QLineEdit(self.centralwidget)
        self.sc_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.sc_line.setObjectName("sc_line")
        self.d_sc_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_sc_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_sc_line.setObjectName("d_sc_line")
        self.u_sc_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_sc_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_sc_line.setObjectName("u_sc_line")
        self.th_line = QtWidgets.QLineEdit(self.centralwidget)
        self.th_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.th_line.setObjectName("th_line")
        self.th_term_line = QtWidgets.QLineEdit(self.centralwidget)
        self.th_term_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.th_term_line.setObjectName("th_term_line")
        self.d_th_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_th_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_th_line.setObjectName("d_th_line")
        self.d_th_term_line = QtWidgets.QLineEdit(self.centralwidget)
        self.d_th_term_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.d_th_term_line.setObjectName("d_th_term_line")
        self.u_th_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_th_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_th_line.setObjectName("u_th_line")
        self.u_th_term_line = QtWidgets.QLineEdit(self.centralwidget)
        self.u_th_term_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.u_th_term_line.setObjectName("u_th_term_line")
        self.ved_start_line =QtWidgets.QLineEdit(self.centralwidget)
        self.ved_start_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_start_line.setObjectName("ved_start_line")
        self.ved_start_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.ved_end_line = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_end_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_end_line.setObjectName("ved_end_line")
        self.ved_end_line.setPlaceholderText('ГГГГ-ММ-ДД')
        self.ved_year_line = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_year_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_year_line.setObjectName("ved_year_line")
        self.ved_umr_line = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_umr_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_umr_line.setObjectName("ved_umr_line")
        self.ved_form_line = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_form_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_form_line.setObjectName("ved_form_line")
        self.ved_term = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_term.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_term.setObjectName("ved_term")
        self.ved_num_group = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_num_group.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_num_group.setObjectName("ved_num_group")
        self.ved_type_line = QtWidgets.QLineEdit(self.centralwidget)
        self.ved_type_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.ved_type_line.setObjectName("ved_type_line")
        self.stud_name = QtWidgets.QLineEdit(self.centralwidget)
        self.stud_name.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.stud_name.setObjectName("stud_name")
        self.stud_num_tiket = QtWidgets.QLineEdit(self.centralwidget)
        self.stud_num_tiket.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.stud_num_tiket.setObjectName("stud_num_tiket")
        self.stud_num_group = QtWidgets.QLineEdit(self.centralwidget)
        self.stud_num_group.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.stud_num_group.setObjectName("stud_num_group")
        self.path_line = QtWidgets.QLineEdit(self.centralwidget)
        self.path_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.path_line.setObjectName("path_line")
        self.login_line = QtWidgets.QLineEdit(self.centralwidget)
        self.login_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.login_line.setObjectName("login_line")
        self.password_line = QtWidgets.QLineEdit(self.centralwidget)
        self.password_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.password_line.setObjectName("password_line")
        self.user_line = QtWidgets.QLineEdit(self.centralwidget)
        self.user_line.setGeometry(QtCore.QRect(10, 130, 221, 25))
        self.user_line.setObjectName("user_line")
        self.add_stud = QtWidgets.QPushButton(self.centralwidget)
        self.add_stud.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.add_stud.setObjectName("add_stud")
        self.add_stud.findChild(QPushButton, 'add_stud')
        self.up_stud = QtWidgets.QPushButton(self.centralwidget)
        self.up_stud.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.up_stud.setObjectName("up_stud")
        self.up_stud.findChild(QPushButton, 'up_stud')
        self.del_stud = QtWidgets.QPushButton(self.centralwidget)
        self.del_stud.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.del_stud.setObjectName("del_stud")
        self.del_stud.findChild(QPushButton, 'del_stud')
        self.add_s = QtWidgets.QPushButton(self.centralwidget)
        self.add_s.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.add_s.setObjectName("add_s")
        self.add_s.findChild(QPushButton, 'add_s')
        self.add_sc = QtWidgets.QPushButton(self.centralwidget)
        self.add_sc.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.add_sc.setObjectName("add_sc")
        self.add_sc.findChild(QPushButton, 'add_sc')
        self.add_th = QtWidgets.QPushButton(self.centralwidget)
        self.add_th.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.add_th.setObjectName("add_th")
        self.add_th.findChild(QPushButton, 'add_th')
        self.add_wd = QtWidgets.QPushButton(self.centralwidget)
        self.add_wd.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.add_wd.setObjectName("add_wd")
        self.add_wd.findChild(QPushButton, 'add_wd')
        self.up_s = QtWidgets.QPushButton(self.centralwidget)
        self.up_s.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.up_s.setObjectName("up_s")
        self.up_s.findChild(QPushButton, 'up_s')
        self.up_sc = QtWidgets.QPushButton(self.centralwidget)
        self.up_sc.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.up_sc.setObjectName("up_sc")
        self.up_sc.findChild(QPushButton, 'up_sc')
        self.up_th = QtWidgets.QPushButton(self.centralwidget)
        self.up_th.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.up_th.setObjectName("up_th")
        self.up_th.findChild(QPushButton, 'up_th')
        self.up_wd = QtWidgets.QPushButton(self.centralwidget)
        self.up_wd.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.up_wd.setObjectName("up_wd")
        self.up_wd.findChild(QPushButton, 'up_wd')
        self.del_s = QtWidgets.QPushButton(self.centralwidget)
        self.del_s.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.del_s.setObjectName("del_s")
        self.del_s.findChild(QPushButton, 'del_s')
        self.del_sc = QtWidgets.QPushButton(self.centralwidget)
        self.del_sc.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.del_sc.setObjectName("del_sc")
        self.del_sc.findChild(QPushButton, 'del_sc')
        self.del_th = QtWidgets.QPushButton(self.centralwidget)
        self.del_th.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.del_th.setObjectName("del_th")
        self.del_th.findChild(QPushButton, 'del_th')
        self.del_wd = QtWidgets.QPushButton(self.centralwidget)
        self.del_wd.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.del_wd.setObjectName("del_wd")
        self.del_wd.findChild(QPushButton, 'del_wd')
        self.gen_ved_btn = QtWidgets.QPushButton(self.centralwidget)
        self.gen_ved_btn.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.gen_ved_btn.setObjectName("gen_ved_btn")
        self.gen_ved_btn.findChild(QPushButton, 'gen_ved_btn')
        self.save_set = QtWidgets.QPushButton(self.centralwidget)
        self.save_set.setGeometry(QtCore.QRect(30, 150, 241, 55))
        self.save_set.setObjectName("save_set")
        self.save_set.findChild(QPushButton, 'save_set')
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 2560, 1600))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "АИС Деканат"))
        self.label_222.setText(_translate("MainWindow", "Введите эл. почту получателя"))
        self.label_221.setText(_translate("MainWindow", "Введите пароль"))
        self.label_220.setText(_translate("MainWindow", "Введите логин"))
        self.label_213.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_212.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_211.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_210.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_203.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_202.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_201.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_200.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_196.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_195.setText(_translate("MainWindow", "Введите номер зачётной книжки"))
        self.label_194.setText(_translate("MainWindow", "Введите ФИО студента"))
        self.label_193.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_192.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_191.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_190.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_183.setText(_translate("MainWindow", "Введите номер семестра"))
        self.label_182.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_181.setText(_translate("MainWindow", "Введите тип мероприятия"))
        self.label_180.setText(_translate("MainWindow", "Введите форму проведения"))
        self.label_179.setText(_translate("MainWindow", "Введите ФИО Зам.директора по УМР"))
        self.label_178.setText(_translate("MainWindow", "Введите год"))
        self.label_177.setText(_translate("MainWindow", "Введите дату сдачи"))
        self.label_176.setText(_translate("MainWindow", "Введите дату выдачи"))
        self.label_175.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_174.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_173.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_172.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_171.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_170.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_161.setText(_translate("MainWindow", "Введите дату:"))
        self.label_160.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_156.setText(_translate("MainWindow", "Введите семестр:"))
        self.label_155.setText(_translate("MainWindow", "Введите кол-во часов:"))
        self.label_154.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_153.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_152.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_151.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_150.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_148.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_147.setText(_translate("MainWindow", "Введите дату"))
        self.label_146.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_145.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_144.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_143.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_142.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_141.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_140.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_131.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_130.setText(_translate("MainWindow", "Выберите расписание:"))
        self.label_129.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_128.setText(_translate("MainWindow", "Выберите название дня недели:"))
        self.label_127.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_126.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_125.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_124.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_123.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_122.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_121.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_111.setText(_translate("MainWindow", "Введите дату:"))
        self.label_110.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_106.setText(_translate("MainWindow", "Введите семестр:"))
        self.label_105.setText(_translate("MainWindow", "Введите кол-во часов:"))
        self.label_104.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_103.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_102.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_101.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_100.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_98.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_97.setText(_translate("MainWindow", "Введите дату"))
        self.label_96.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_95.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_94.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_93.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_92.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_91.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_90.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_81.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_80.setText(_translate("MainWindow", "Выберите расписание:"))
        self.label_79.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_78.setText(_translate("MainWindow", "Выберите название дня недели:"))
        self.label_77.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_76.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_75.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_74.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_73.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_72.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_71.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_61.setText(_translate("MainWindow", "Введите дату:"))
        self.label_60.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_56.setText(_translate("MainWindow", "Введите семестр:"))
        self.label_55.setText(_translate("MainWindow", "Введите кол-во часов:"))
        self.label_54.setText(_translate("MainWindow", "Выберите организацию:"))
        self.label_53.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_52.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_51.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_50.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_48.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_47.setText(_translate("MainWindow", "Введите дату"))
        self.label_46.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_45.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_44.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_43.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_42.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_41.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_40.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_31.setText(_translate("MainWindow", "Введите номер группы"))
        self.label_30.setText(_translate("MainWindow", "Выберите расписание:"))
        self.label_29.setText(_translate("MainWindow", "Выберите тип недели:"))
        self.label_28.setText(_translate("MainWindow", "Выберите название дня недели:"))
        self.label_27.setText(_translate("MainWindow", "Выберите номер пары:"))
        self.label_26.setText(_translate("MainWindow", "Выберите год поступления:"))
        self.label_25.setText(_translate("MainWindow", "Выберите курс:"))
        self.label_24.setText(_translate("MainWindow", "Выберите группу:"))
        self.label_23.setText(_translate("MainWindow", "Выберите преподавателя:"))
        self.label_22.setText(_translate("MainWindow", "Выберите образовательную организацию:"))
        self.label_21.setText(_translate("MainWindow", "Выберите дисциплину:"))
        self.label_6.setText(_translate("MainWindow", "Укажите путь к папке:"))
        self.label_5.setText(_translate("MainWindow", "Укажите путь к папке:"))
        self.label_4.setText(_translate("MainWindow", "Укажите путь к папке:"))
        self.label_3.setText(_translate("MainWindow", "Укажите путь к папке:"))
        self.label_2.setText(_translate("MainWindow", "Укажите путь к папке:"))
        self.label1.setText(_translate("MainWindow", "Укажите путь к файлу:"))
        self.label.setText(_translate("MainWindow", "Укажите путь к файлу:"))
        self.open_pars_date_week.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_s.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_sc.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_th.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_sub.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_tea.setText(_translate("MainWindow", "Обзор..."))
        self.open_pars_stud.setText(_translate("MainWindow", "Обзор..."))
        self.pars_date_week.setText(_translate("MainWindow", "Загрузить"))
        self.pars_s.setText(_translate("MainWindow", "Загрузить"))
        self.pars_sc.setText(_translate("MainWindow", "Загрузить"))
        self.pars_th.setText(_translate("MainWindow", "Загрузить"))
        self.pars_sub.setText(_translate("MainWindow", "Загрузить"))
        self.pars_tea.setText(_translate("MainWindow", "Загрузить"))
        self.pars_stud.setText(_translate("MainWindow", "Загрузить"))
        self.add_s.setText(_translate("MainWindow", "Сохранить"))
        self.add_sc.setText(_translate("MainWindow", "Сохранить"))
        self.add_th.setText(_translate("MainWindow", "Сохранить"))
        self.add_wd.setText(_translate("MainWindow", "Сохранить"))
        self.up_s.setText(_translate("MainWindow", "Открыть"))
        self.up_sc.setText(_translate("MainWindow", "Открыть"))
        self.up_th.setText(_translate("MainWindow", "Открыть"))
        self.up_wd.setText(_translate("MainWindow", "Открыть"))
        self.del_s.setText(_translate("MainWindow", "Открыть"))
        self.del_sc.setText(_translate("MainWindow", "Открыть"))
        self.del_th.setText(_translate("MainWindow", "Открыть"))
        self.del_wd.setText(_translate("MainWindow", "Открыть"))
        self.gen_ved_btn.setText(_translate("MainWindow", "Генерировать"))
        self.add_stud.setText(_translate("MainWindow", "Сохранить"))
        self.up_stud.setText(_translate("MainWindow", "Открыть"))
        self.del_stud.setText(_translate("MainWindow", "Открыть"))
        self.save_set.setText(_translate("MainWindow", "Сохранить"))


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    customFormat = 'yyyy-MM-dd'
    def __init__(self, login):
        sel_test = 'SELECT * FROM `users` WHERE `login` = %s'
        l = []
        l.append(login)
        cursor.execute(sel_test, l)
        self.user = cursor.fetchone()
        super().__init__()
        self.showFullScreen()
        self.setupUi(self)
        self.parser_date_week_ = QWidget()
        self.parser_s = QWidget()
        self.parser_sc = QWidget()
        self.parser_th = QWidget()
        self.addSchedule = QWidget()
        self.addScheduleChange = QWidget()
        self.addTheme = QWidget()
        self.addWeekDate = QWidget()
        self.updateSchedule = QWidget()
        self.updateScheduleChange = QWidget()
        self.updateTheme = QWidget()
        self.updateWeekDate = QWidget()
        self.deleteSchedule = QWidget()
        self.deleteScheduleChange = QWidget()
        self.deleteTheme = QWidget()
        self.deleteWeekDate = QWidget()
        self.gen_ved_ = QWidget()
        self.pars_sub_ = QWidget()
        self.pars_tea_ = QWidget()
        self.pars_stud_ = QWidget()
        self.addStud = QWidget()
        self.updateStud = QWidget()
        self.deleteStud = QWidget()
        self.set_email_ = QWidget()
        self.tabWidget = QTabWidget(self.centralwidget)
        self.tabWidget.addTab(self.parser_date_week_, "Загрузка даты типов недели")
        self.tabWidget.addTab(self.parser_s, "Загрузка основного расписания")
        self.tabWidget.addTab(self.parser_sc, "Загрузка замен в расписании")
        self.tabWidget.addTab(self.parser_th, "Загрузка часов тем")
        self.tabWidget.addTab(self.addSchedule, "Добавить пункт в расписание")
        self.tabWidget.addTab(self.addScheduleChange, "Добавить пункт в замен расписания")
        self.tabWidget.addTab(self.addTheme, "Добавить тему в план")
        self.tabWidget.addTab(self.addWeekDate, "Добавить дату дня")
        self.tabWidget.addTab(self.updateSchedule, "Обновить пункт расписания")
        self.tabWidget.addTab(self.updateScheduleChange, "Обновить замену в расписании")
        self.tabWidget.addTab(self.updateTheme, "Обновить тему в плане")
        self.tabWidget.addTab(self.updateWeekDate, "Обновить дату дня")
        self.tabWidget.addTab(self.deleteSchedule, "Удалить пункт из расписания")
        self.tabWidget.addTab(self.deleteScheduleChange, "Удалить замену в расписании")
        self.tabWidget.addTab(self.deleteTheme, "Удалить тему из плана")
        self.tabWidget.addTab(self.deleteWeekDate, "Удалить дату дня")
        self.tabWidget.addTab(self.gen_ved_, "Генератор экзаменационной ведомости")
        self.tabWidget.addTab(self.pars_sub_, "Загрузка дисциплин")
        self.tabWidget.addTab(self.pars_tea_, "Загрузка преподавателей")
        self.tabWidget.addTab(self.pars_stud_, "Загрузка студентов")
        self.tabWidget.addTab(self.addStud, "Добавить студентов")
        self.tabWidget.addTab(self.updateStud, "Обновить студентов")
        self.tabWidget.addTab(self.deleteStud, "Удалить студентов")
        self.tabWidget.addTab(self.set_email_, "Настройка почтового адреса")
        self.create_schedule_Ui()
        self.create_schedule_change_Ui()
        self.create_date_type_week_Ui()
        self.update_schedule_changes_Ui()
        self.update_date_week_Ui()
        self.delete_schedule_changes_Ui()
        self.parser_schedule_Ui()
        self.parser_schedule_changes_Ui()
        self.parser_th_Ui()
        self.parser_date_week_Ui()
        self.create_theme_hours_Ui()
        self.update_schedule_Ui()
        self.update_theme_hours_Ui()
        self.delete_schedule_Ui()
        self.delete_theme_hours_Ui()
        self.delete_date_week_Ui()
        self.parser_subjects_Ui()
        self.parser_teachers_Ui()
        self.parser_students_Ui()
        self.create_students_Ui()
        self.update_students_Ui()
        self.delete_students_Ui()
        self.set_email_Ui()
        self.open_pars_date_week.clicked.connect(self.pars_win_date_week)
        self.pars_date_week.clicked.connect(self.parser_date_week)
        self.open_pars_s.clicked.connect(self.pars_win_schedule)
        self.open_pars_sc.clicked.connect(self.pars_win_schedule_changes)
        self.open_pars_th.clicked.connect(self.pars_win_th)
        self.open_pars_sub.clicked.connect(self.pars_win_sub)
        self.open_pars_tea.clicked.connect(self.pars_win_tea)
        self.open_pars_stud.clicked.connect(self.pars_win_stud)
        self.pars_s.clicked.connect(self.parser_schedule)
        self.pars_sc.clicked.connect(self.parser_schedule_changes)
        self.pars_th.clicked.connect(self.parser_theme_h)
        self.pars_sub.clicked.connect(self.parser_subjects)
        self.pars_tea.clicked.connect(self.parser_teachers)
        self.pars_stud.clicked.connect(self.parser_students)
        self.add_s.clicked.connect(self.create_schedule)
        self.add_sc.clicked.connect(self.create_schedule_change)
        self.add_th.clicked.connect(self.create_theme_hours)
        self.add_wd.clicked.connect(self.create_date_type_week)
        self.add_stud.clicked.connect(self.create_students)
        self.up_s.clicked.connect(self.update_schedule)
        self.up_sc.clicked.connect(self.update_schedule_changes)
        self.up_th.clicked.connect(self.update_theme_hours)
        self.up_wd.clicked.connect(self.update_date_week)
        self.up_stud.clicked.connect(self.update_students)
        self.del_s.clicked.connect(self.delete_schedule)
        self.del_sc.clicked.connect(self.delete_schedule_changes)
        self.del_th.clicked.connect(self.delete_theme_hours)
        self.del_wd.clicked.connect(self.delete_date_week)
        self.del_stud.clicked.connect(self.delete_students)
        self.gen_ved_btn.clicked.connect(self.gen_ved)
        self.save_set.clicked.connect(self.set_email)

    def set_email_Ui(self):
        self.tabWidget = QTabWidget(self.centralwidget)
        vl = QVBoxLayout(self)
        vl.addWidget(self.label_220)
        vl.addWidget(self.login_line)
        vl.addWidget(self.label_221)
        vl.addWidget(self.password_line)
        vl.addWidget(self.label_222)
        vl.addWidget(self.user_line)
        vl.addWidget(self.save_set)
        self.tabWidget.setTabText(23, "SetEmail")
        self.set_email_.setLayout(vl)

    def check_save_set_email(self):
        return os.path.exists('config_set_email.json')

    def set_email(self):
        if self.check_save_set_email():
            data = [{'login': self.login_line.text(), 'password': self.password_line.text(), 'user': self.user_line.text()}]
            with open('config_set_email.json', 'w') as save:
                json.dump(data, save)
        else:
            workbook = Workbook()
            worksheet = workbook.worksheets[0]
            worksheet.cells.get("A1").put_value("login")
            worksheet.cells.get("B1").put_value("password")
            worksheet.cells.get("C1").put_value("user")
            worksheet.cells.get("A2").put_value(self.login_line.text())
            worksheet.cells.get("B2").put_value(self.password_line.text())
            worksheet.cells.get("C2").put_value(self.user_line.text())
            workbook.save("config_set_email.json")


    def gen_ved_Ui(self):
        self.combo90 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo90.addItem(check_sel[i][0])
        self.combo91 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo91.addItem(check_sel[i][0])
        self.combo92 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo92.addItem(check_sel[i][0])
        self.combo93 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo93.addItem(check_sel[i][0])
        self.combo94 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo94.addItem(check_sel[i][0])
        self.combo95 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo95.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_170)
        vlayout.addWidget(self.combo90)
        vlayout.addWidget(self.label_171)
        vlayout.addWidget(self.combo91)
        vlayout.addWidget(self.label_172)
        vlayout.addWidget(self.combo92)
        vlayout.addWidget(self.label_173)
        vlayout.addWidget(self.combo93)
        vlayout.addWidget(self.label_174)
        vlayout.addWidget(self.combo94)
        vlayout.addWidget(self.label_175)
        vlayout.addWidget(self.combo95)
        vlayout.addWidget(self.label_176)
        vlayout.addWidget(self.ved_start_line)
        vlayout.addWidget(self.label_177)
        vlayout.addWidget(self.ved_end_line)
        vlayout.addWidget(self.label_178)
        vlayout.addWidget(self.ved_year_line)
        vlayout.addWidget(self.label_179)
        vlayout.addWidget(self.ved_umr_line)
        vlayout.addWidget(self.label_180)
        vlayout.addWidget(self.ved_form_line)
        vlayout.addWidget(self.label_181)
        vlayout.addWidget(self.ved_type_line)
        vlayout.addWidget(self.label_182)
        vlayout.addWidget(self.ved_num_group)
        vlayout.addWidget(self.label_183)
        vlayout.addWidget(self.ved_term)
        vlayout.addWidget(self.gen_ved_btn)
        self.tabWidget.setTabText(16, "GenVed")
        self.gen_ved_.setLayout(vlayout)

    def gen_ved(self):
        data_id = []
        gr_name = []
        cour_name = []
        ye_name = []
        org_name = []
        gr_name.append(self.combo92.currentText())
        cour_name.append(self.combo93.currentText())
        ye_name.append(self.combo94.currentText())
        org_name.append(self.combo95.currentText())
        sel_id = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        cursor.execute(sel_id, gr_name)
        data_id.append(cursor.fetchone()[0])
        sel_id = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        cursor.execute(sel_id, cour_name)
        data_id.append(cursor.fetchone()[0])
        sel_id = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        cursor.execute(sel_id, ye_name)
        data_id.append(cursor.fetchone()[0])
        sel_id = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_id, org_name)
        data_id.append(cursor.fetchone()[0])
        data_id.append(int(self.ved_num_group.text()))
        sel_students = 'SELECT `name`, `number_stud_tiket` FROM `students` WHERE `groups_id` = %s AND `courses_id` = %s AND `year_enter_id` = %s ' \
                       'AND `organization_id` = %s AND `num_group` = %s ORDER BY `name`'
        cursor.execute(sel_students, data_id)
        arr_students = cursor.fetchall()
        doc = aw.Document()
        builder = aw.DocumentBuilder(doc)
        builder.start_table()
        builder.insert_cell()
        imageAsLinkToFile = builder.insert_image("logo.png")
        imageAsLinkToFile.width = aw.ConvertUtil.inch_to_point(1.6)
        imageAsLinkToFile.height = aw.ConvertUtil.inch_to_point(1)
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 14
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.cell_format.width = 300.0
        builder.write("АНПОО «Тамбовский колледж бизнес-технологий»")
        builder.end_row()
        builder.end_table()
        builder.write('________________________________________________________________________________')
        builder.write('\n')
        builder.write('\n')
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write(
            "392020, Тамбов, ул. Пензенская/Карла Маркса, д.61/175, корпус 3\nТел.: (4752) 77-10-64 E-mail: mail@tkbt68.ru")
        builder.write('\n')
        builder.write('\n')
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 14
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write('ЗАЧЕТНО-ЭКЗАМЕНАЦИОННАЯ ВЕДОМОСТЬ')
        builder.write('\n')
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write('№  / _____  от  / «___» _________ ' + str(self.ved_year_line.text()) + ' г.')
        builder.write('\n')
        builder.write('\n')
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write('Преподаватели ')
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write(self.combo91.currentText())
        builder.start_table()
        builder.insert_cell()
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Группа")
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write("к"+str(self.combo92.currentText())+"-"+str(self.combo94.currentText())+"/"+ str(self.ved_num_group.text()) + " группа, "+str(self.combo93.currentText())+ " курс")
        builder.end_row()
        builder.insert_cell()
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Учебный период")
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write(str(self.ved_term.text()) + " семестр")
        builder.end_row()
        builder.insert_cell()
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Дисциплина")
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write(str(self.combo90.currentText()))
        builder.end_row()
        builder.insert_cell()
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Форма проведения")
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.cell_format.width = 300.0
        builder.write(self.ved_form_line.text())
        builder.end_row()
        builder.end_table()
        builder.write('\n')
        builder.start_table()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 11
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Дата выдачи")
        builder.insert_cell()
        builder.write("Дата сдачи")
        builder.insert_cell()
        builder.write("Вид мероприятия")
        builder.end_row()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 11
        builder.font.name = "Times New Roman"
        builder.font.bold = True
        builder.write(str(self.ved_start_line.text()))
        builder.insert_cell()
        builder.write(str(self.ved_end_line.text()))
        builder.insert_cell()
        builder.write("Зачет")
        builder.end_row()
        builder.end_table()
        builder.write('\n')
        builder.start_table()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
        builder.font.size = 11
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("№ п/п")
        builder.insert_cell()
        builder.write("Фамилия и инициалы")
        builder.insert_cell()
        builder.write("№ зачетной книжки")
        builder.insert_cell()
        builder.write("Оценка")
        builder.insert_cell()
        builder.write("Подпись")
        builder.end_row()
        for i in range(0, len(arr_students)):
            for j in range(0, len(arr_students[i])):
                builder.insert_cell()
                builder.paragraph_format.alignment = aw.ParagraphAlignment.CENTER
                builder.font.size = 12
                builder.font.name = "Times New Roman"
                builder.font.bold = False
                builder.write(i+1)
                builder.insert_cell()
                builder.write(str(arr_students[i][0]))
                builder.insert_cell()
                builder.cell_format.width = 300.0
                builder.write(str(arr_students[i][1]))
                builder.cell_format.width = 300.0
                builder.insert_cell()
                builder.write("")
                builder.cell_format.width = 300.0
                builder.insert_cell()
                builder.write("")
                builder.end_row()
        builder.end_table()
        builder.write('\n')
        builder.start_table()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Зам.директора по УМР")
        builder.insert_cell()
        builder.write("___________________")
        builder.insert_cell()
        builder.write("// " + str(self.ved_umr_line.text()))
        builder.end_row()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 12
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("Преподаватель")
        builder.insert_cell()
        builder.write("___________________")
        builder.insert_cell()
        builder.write("// " + str(self.combo91.currentText()))
        builder.end_row()
        builder.end_table()
        builder.write('\n')
        builder.start_table()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 9
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write("всего")
        builder.insert_cell()
        builder.write("незачтено")
        builder.insert_cell()
        builder.write("неудовлетворительно")
        builder.insert_cell()
        builder.write("удовлетворительно")
        builder.insert_cell()
        builder.write("хорошо")
        builder.insert_cell()
        builder.write("отлично")
        builder.insert_cell()
        builder.write("зачтено")
        builder.insert_cell()
        builder.write("на проверке")
        builder.insert_cell()
        builder.write("не аттест.")
        builder.insert_cell()
        builder.write("неявка")
        builder.end_row()
        builder.insert_cell()
        builder.paragraph_format.alignment = aw.ParagraphAlignment.LEFT
        builder.font.size = 9
        builder.font.name = "Times New Roman"
        builder.font.bold = False
        builder.write(str(len(arr_students)))
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.insert_cell()
        builder.write("")
        builder.end_table()
        doc.save("к"+str(self.combo92.currentText())+"-"+str(self.combo94.currentText())+"/"+ str(self.ved_num_group.text()) + " группа, "+str(self.combo93.currentText())+ " курс "+ str(self.combo90.currentText()) + " " + str(self.combo91.currentText())+ ".docx")

    def parser_subjects_Ui(self):
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label_4)
        hlayout.addWidget(self.filename_4)
        hlayout.addWidget(self.open_pars_sub)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_sub)
        self.tabWidget.setTabText(17, "ParserSub")
        self.pars_sub_.setLayout(vlayout)

    def parser_teachers_Ui(self):
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label_5)
        hlayout.addWidget(self.filename_5)
        hlayout.addWidget(self.open_pars_tea)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_tea)
        self.tabWidget.setTabText(18, "ParserTea")
        self.pars_tea_.setLayout(vlayout)

    def parser_students_Ui(self):
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label_6)
        hlayout.addWidget(self.filename_6)
        hlayout.addWidget(self.open_pars_stud)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_stud)
        self.tabWidget.setTabText(19, "ParserStud")
        self.pars_stud_.setLayout(vlayout)

    def parser_students(self):
        df = pd.read_excel(io=self.file_, engine='openpyxl', sheet_name='Лист1')
        arr = list(df.head(0))

        # Парс excel
        result = []

        # Парс ОРГАНИЗАЦИИ
        for i in range(0, len(df[arr[0]].tolist())):
            add_ser = 'INSERT INTO `organization` (`name`) VALUES (%s)'
            result.append(df[arr[0]].tolist()[i])
            check_input = 'SELECT * FROM `organization` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГРУПП
        for i in range(0, len(df[arr[3]].tolist())):
            add_ser = 'INSERT INTO `groups` (`name`) VALUES (%s)'
            result.append(df[arr[3]].tolist()[i])
            check_input = 'SELECT * FROM `groups` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс КУРСОВ
        for i in range(0, len(df[arr[4]].tolist())):
            add_ser = 'INSERT INTO `courses` (`name`) VALUES (%s)'
            result.append(df[arr[4]].tolist()[i])
            check_input = 'SELECT * FROM `courses` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГОД ПОСТУПЛЕНИЙ
        for i in range(0, len(df[arr[6]].tolist())):
            add_ser = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
            result.append(df[arr[6]].tolist()[i])
            check_input = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс СТУДЕНТОВ
        for i in range(0, len(df[arr[0]].tolist())):
            add_ser = 'INSERT INTO `students` (`organization_id`, `name`, `number_stud_tiket`, `groups_id`, ' \
                      '`courses_id`, `num_group`, `year_enter_id`) VALUES (%s, %s, %s, %s, %s, %s, %s)'
            result = []
            for j in range(0, len(arr)):
                data_db = []
                if j == 0:
                    data_db.append(self.combo70.currentText())
                    check_input = 'SELECT `id` FROM `organization` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(int(cursor.fetchone()[0]))
                if j == 1:
                    result.append(df[arr[j]][i])
                if j == 2:
                    result.append(str(df[arr[j]][i]))
                if j == 3:
                    data_db.append(df[arr[j]][i])
                    check_input = 'SELECT `id` FROM `groups` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(int(cursor.fetchone()[0]))
                if j == 4:
                    data_db.append(str(df[arr[j]][i]))
                    check_input = 'SELECT `id` FROM `courses` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(int(cursor.fetchone()[0]))
                if j == 5:
                    result.append(int(df[arr[j]][i]))
                if j == 6:
                    data_db.append(str(df[arr[j]][i]))
                    check_input = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(int(cursor.fetchone()[0]))
                    check_input = 'SELECT * FROM `students` WHERE `organization_id` = %s AND `name` = %s AND ' \
                                  '`number_stud_tiket` = %s AND `groups_id` = %s AND `courses_id` = %s AND `num_group` = %s ' \
                                  'AND `year_enter_id` = %s'
                    cursor.execute(check_input, result)
                    if cursor.fetchone() == None:
                        cursor.execute(add_ser, result)
                        conn.commit()
                        result = []
                    else:
                        result = []

    def parser_teachers(self):
        df = pd.read_excel(io=self.file_, engine='openpyxl', sheet_name='Лист1')
        arr = list(df.head(0))

        # Парс excel
        result = []

        # Парс ОРГАНИЗАЦИИ
        for i in range(0, len(df[arr[0]].tolist())):
            add_ser = 'INSERT INTO `organization` (`name`) VALUES (%s)'
            result.append(df[arr[0]].tolist()[i])
            check_input = 'SELECT * FROM `organization` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ДИСЦИПЛИН
        for i in range(0, len(df[arr[1]].tolist())):
            add_ser = 'INSERT INTO `teachers` (`name`) VALUES (%s)'
            result.append(df[arr[1]].tolist()[i])
            check_input = 'SELECT * FROM `teachers` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

    def parser_subjects(self):
        df = pd.read_excel(io=self.file_, engine='openpyxl', sheet_name='Лист1')
        arr = list(df.head(0))

        # Парс excel
        result = []

        # Парс ОРГАНИЗАЦИИ
        for i in range(0, len(df[arr[0]].tolist())):
            add_ser = 'INSERT INTO `organization` (`name`) VALUES (%s)'
            result.append(df[arr[0]].tolist()[i])
            check_input = 'SELECT * FROM `organization` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ДИСЦИПЛИН
        for i in range(0, len(df[arr[1]].tolist())):
            add_ser = 'INSERT INTO `subjects` (`name`) VALUES (%s)'
            result.append(df[arr[1]].tolist()[i])
            check_input = 'SELECT * FROM `subjects` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

    def pars_win_sub(self):
        self.open_pars_sub.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename_4.setText(fname)
            self.file_ = str(fname)

    def pars_win_tea(self):
        self.open_pars_tea.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename_5.setText(fname)
            self.file_ = str(fname)

    def pars_win_stud(self):
        self.open_pars_stud.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename_6.setText(fname)
            self.file_ = str(fname)

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_F11:
            # если в полный экран
            if self.isFullScreen():
                # вернуть прежнее состояние
                self.showNormal()
            else:
                # иначе во весь экран
                self.showFullScreen()
        super().keyPressEvent(event)

    def create_students_Ui(self):
        self.combo200 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo200.addItem(check_sel[i][0])
        self.combo201 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo201.addItem(check_sel[i][0])
        self.combo202 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo202.addItem(check_sel[i][0])
        self.combo203 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo203.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_190)
        vlayout.addWidget(self.combo200)
        vlayout.addWidget(self.label_191)
        vlayout.addWidget(self.combo201)
        vlayout.addWidget(self.label_192)
        vlayout.addWidget(self.combo202)
        vlayout.addWidget(self.label_193)
        vlayout.addWidget(self.combo203)
        vlayout.addWidget(self.label_194)
        vlayout.addWidget(self.stud_name)
        vlayout.addWidget(self.label_195)
        vlayout.addWidget(self.stud_num_tiket)
        vlayout.addWidget(self.label_196)
        vlayout.addWidget(self.stud_num_group)
        vlayout.addWidget(self.add_stud)
        self.tabWidget.setTabText(20, "CteateStud")
        self.addStud.setLayout(vlayout)

    def create_students(self):
        data = []
        data.append(self.combo200.currentText())
        data.append(self.combo201.currentText())
        data.append(self.combo202.currentText())
        data.append(self.combo203.currentText())
        data.append(self.stud_name.text())
        data.append(self.stud_num_tiket.text())
        data.append(self.stud_num_group.text())
        data_id = []
        name_gr = []
        name_cour = []
        name_ye = []
        name_org = []
        name_gr.append(data[0])
        name_cour.append(data[1])
        name_ye.append(data[2])
        name_org.append(data[3])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_gr, name_gr)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_cour, name_cour)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_ye, name_ye)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_org, name_org)
        data_id.append(cursor.fetchone()[0])
        data_id.append(data[4])
        data_id.append(data[5])
        data_id.append(data[6])
        in_schedule = 'INSERT INTO `students` (`groups_id`, `courses_id`, `year_enter_id`, `organization_id`, ' \
                      '`name`, `number_stud_tiket`, `num_group`) ' \
                      'VALUES (%s, %s, %s, %s, %s, %s, %s)'
        check_in = 'SELECT * FROM `students` WHERE `groups_id` = %s AND `courses_id` = %s AND `year_enter_id` = %s ' \
                   'AND `organization_id` = %s AND `name` = %s AND `number_stud_tiket` = %s AND `num_group` = %s'
        cursor.execute(check_in, data_id)
        if cursor.fetchone() == None:
            cursor.execute(in_schedule, data_id)
            conn.commit()

    def update_students_Ui(self):
        self.combo200 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo200.addItem(check_sel[i][0])
        self.combo201 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo201.addItem(check_sel[i][0])
        self.combo202 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo202.addItem(check_sel[i][0])
        self.combo203 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo203.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_200)
        vlayout.addWidget(self.combo200)
        vlayout.addWidget(self.label_201)
        vlayout.addWidget(self.combo201)
        vlayout.addWidget(self.label_202)
        vlayout.addWidget(self.combo202)
        vlayout.addWidget(self.label_203)
        vlayout.addWidget(self.combo203)
        vlayout.addWidget(self.up_stud)
        self.tabWidget.setTabText(21, "UpdateStud")
        self.updateStud.setLayout(vlayout)

    def update_students(self):
        students = []
        students.append(self.combo200.currentText())
        students.append(self.combo201.currentText())
        students.append(self.combo202.currentText())
        students.append(self.combo203.currentText())
        self.UStud = UpdateStudents(students)
        self.UStud.show()

    def delete_students_Ui(self):
        self.combo200 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo200.addItem(check_sel[i][0])
        self.combo201 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo201.addItem(check_sel[i][0])
        self.combo202 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo202.addItem(check_sel[i][0])
        self.combo203 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo203.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_210)
        vlayout.addWidget(self.combo200)
        vlayout.addWidget(self.label_211)
        vlayout.addWidget(self.combo201)
        vlayout.addWidget(self.label_212)
        vlayout.addWidget(self.combo202)
        vlayout.addWidget(self.label_213)
        vlayout.addWidget(self.combo203)
        vlayout.addWidget(self.del_stud)
        self.tabWidget.setTabText(22, "DeleteStud")
        self.deleteStud.setLayout(vlayout)

    def delete_students(self):
        students = []
        students.append(self.combo200.currentText())
        students.append(self.combo201.currentText())
        students.append(self.combo202.currentText())
        students.append(self.combo203.currentText())
        self.DStud = DeleteStudent(students)
        self.DStud.show()

    def create_schedule_Ui(self):
        self.combo30 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo30.addItem(check_sel[i][0])
        self.combo31 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo31.addItem(check_sel[i][0])
        self.combo32 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo32.addItem(check_sel[i][0])
        self.combo33 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo33.addItem(check_sel[i][0])
        self.combo34 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo34.addItem(check_sel[i][0])
        self.combo35 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo35.addItem(check_sel[i][0])
        self.combo36 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `name_day`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo36.addItem(check_sel[i][0])
        self.combo37 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo37.addItem(check_sel[i][0])
        self.combo38 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo38.addItem(check_sel[i][0])
        self.combo39 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `sprav_schedule`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo39.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_21)
        vlayout.addWidget(self.combo30)
        vlayout.addWidget(self.label_23)
        vlayout.addWidget(self.combo31)
        vlayout.addWidget(self.label_24)
        vlayout.addWidget(self.combo32)
        vlayout.addWidget(self.label_25)
        vlayout.addWidget(self.combo33)
        vlayout.addWidget(self.label_26)
        vlayout.addWidget(self.combo34)
        vlayout.addWidget(self.label_27)
        vlayout.addWidget(self.combo35)
        vlayout.addWidget(self.label_28)
        vlayout.addWidget(self.combo36)
        vlayout.addWidget(self.label_29)
        vlayout.addWidget(self.combo37)
        vlayout.addWidget(self.label_22)
        vlayout.addWidget(self.combo38)
        vlayout.addWidget(self.label_30)
        vlayout.addWidget(self.combo39)
        vlayout.addWidget(self.label_31)
        vlayout.addWidget(self.schedule_line)
        vlayout.addWidget(self.add_s)
        self.tabWidget.setTabText(4, "CteateS")
        self.addSchedule.setLayout(vlayout)

    def create_schedule(self):
        data = []
        data.append(self.combo30.currentText())
        data.append(self.combo31.currentText())
        data.append(self.combo32.currentText())
        data.append(self.combo33.currentText())
        data.append(self.combo34.currentText())
        data.append(self.combo35.currentText())
        data.append(self.combo36.currentText())
        data.append(self.combo37.currentText())
        data.append(self.combo38.currentText())
        data.append(self.combo39.currentText())
        data.append(self.schedule_line.text())
        data_id = []
        name_sub = []
        name_tea = []
        name_gr = []
        name_cour = []
        name_ye = []
        name_nl = []
        name_nd = []
        name_tw = []
        name_org = []
        name_ss = []
        name_sub.append(data[0])
        name_tea.append(data[1])
        name_gr.append(data[2])
        name_cour.append(data[3])
        name_ye.append(data[4])
        name_nl.append(data[5])
        name_nd.append(data[6])
        name_tw.append(data[7])
        name_org.append(data[8])
        name_ss.append(data[9])
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        sel_nl = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        sel_nd = 'SELECT `id` FROM `name_day` WHERE `name` = %s'
        sel_tw = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        sel_ss = 'SELECT `id` FROM `sprav_schedule` WHERE `name` = %s'
        cursor.execute(sel_sub, name_sub)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_tea, name_tea)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_gr, name_gr)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_cour, name_cour)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_ye, name_ye)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_nl, name_nl)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_nd, name_nd)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_tw, name_tw)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_org, name_org)
        data_id.append(cursor.fetchone()[0])
        data_id.append(data[10])
        cursor.execute(sel_ss, name_ss)
        data_id.append(cursor.fetchone()[0])
        in_schedule = 'INSERT INTO `schedule` (`subjects_id`, `teachers_id`, `groups_id`, `courses_id`, `year_enter_id`, ' \
                   '`num_lessons_id`, `name_day_id`, `type_week_id`, `organization_id`, `num_group`, `sprav_schedule_id`) ' \
                   'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
        check_in = 'SELECT * FROM `schedule` WHERE `subjects_id` = %s AND `teachers_id` = %s AND `groups_id` = %s AND ' \
                   '`courses_id` = %s AND `year_enter_id` = %s AND `num_lessons_id` = %s AND `name_day_id` = %s AND ' \
                   '`type_week_id` = %s AND `organization_id` = %s AND `num_group` = %s AND `sprav_schedule_id` = %s'
        cursor.execute(check_in, data_id)
        if cursor.fetchone() == None:
            cursor.execute(in_schedule, data_id)
            conn.commit()

    def create_schedule_change_Ui(self):
        self.combo40 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo40.addItem(check_sel[i][0])
        self.combo41 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo41.addItem(check_sel[i][0])
        self.combo42 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo42.addItem(check_sel[i][0])
        self.combo43 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo43.addItem(check_sel[i][0])
        self.combo44 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo44.addItem(check_sel[i][0])
        self.combo45 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo45.addItem(check_sel[i][0])
        self.combo46 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo46.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_40)
        vlayout.addWidget(self.combo40)
        vlayout.addWidget(self.label_41)
        vlayout.addWidget(self.combo41)
        vlayout.addWidget(self.label_42)
        vlayout.addWidget(self.combo42)
        vlayout.addWidget(self.label_43)
        vlayout.addWidget(self.combo43)
        vlayout.addWidget(self.label_44)
        vlayout.addWidget(self.combo44)
        vlayout.addWidget(self.label_45)
        vlayout.addWidget(self.combo45)
        vlayout.addWidget(self.label_46)
        vlayout.addWidget(self.combo46)
        vlayout.addWidget(self.label_47)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.sc_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.label_48)
        vlayout.addWidget(self.sc_line)
        vlayout.addWidget(self.add_sc)
        self.tabWidget.setTabText(5, "CteateSC")
        self.addScheduleChange.setLayout(vlayout)

    def create_schedule_change(self):
        data = []
        data.append(self.combo40.currentText())
        data.append(self.combo41.currentText())
        data.append(self.combo42.currentText())
        data.append(self.combo43.currentText())
        data.append(self.combo44.currentText())
        data.append(self.combo45.currentText())
        data.append(self.combo46.currentText())
        data_id = []
        name_sub = []
        name_tea = []
        name_gr = []
        name_cour = []
        name_ye = []
        name_nl = []
        name_org = []
        name_sub.append(data[0])
        name_tea.append(data[1])
        name_gr.append(data[2])
        name_cour.append(data[3])
        name_ye.append(data[4])
        name_nl.append(data[5])
        name_org.append(data[6])
        if name_sub[0] == '<<Не определено>>' and name_tea[0] == '<<Не определено>>':
            data_id.append(None)
            data_id.append(None)
        else:
            sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
            sel_tea = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
            cursor.execute(sel_sub, name_sub)
            data_id.append(cursor.fetchone()[0])
            cursor.execute(sel_tea, name_tea)
            data_id.append(cursor.fetchone()[0])
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        sel_nl = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_gr, name_gr)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_cour, name_cour)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_ye, name_ye)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_nl, name_nl)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_org, name_org)
        data_id.append(cursor.fetchone()[0])
        data_id.append(self.sc_date_line.text())
        data_id.append(int(self.sc_line.text()))
        in_schedule_change = 'INSERT INTO `schedule_changes` (`subjects_id`, `teachers_id`, `groups_id`, `courses_id`, `year_enter_id`, ' \
                      '`num_lessons_id`, `organization_id`, `date_changes`, `num_group`) ' \
                      'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)'
        check_in = 'SELECT * FROM `schedule_changes` WHERE `subjects_id` = %s AND `teachers_id` = %s AND `groups_id` = %s AND ' \
                   '`courses_id` = %s AND `year_enter_id` = %s AND `num_lessons_id` = %s AND`organization_id` = ' \
                   '%s AND `date_changes` = %s AND `num_group` = %s'
        cursor.execute(check_in, data_id)
        if cursor.fetchone() == None:
            cursor.execute(in_schedule_change, data_id)
            conn.commit()

    def create_theme_hours_Ui(self):
        self.combo50 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo50.addItem(check_sel[i][0])
        self.combo51 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo51.addItem(check_sel[i][0])
        self.combo52 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo52.addItem(check_sel[i][0])
        self.combo53 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo53.addItem(check_sel[i][0])
        self.combo54 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo54.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_50)
        vlayout.addWidget(self.combo50)
        vlayout.addWidget(self.label_51)
        vlayout.addWidget(self.combo51)
        vlayout.addWidget(self.label_52)
        vlayout.addWidget(self.combo52)
        vlayout.addWidget(self.label_53)
        vlayout.addWidget(self.combo53)
        vlayout.addWidget(self.label_54)
        vlayout.addWidget(self.combo54)
        vlayout.addWidget(self.label_55)
        vlayout.addWidget(self.th_line)
        vlayout.addWidget(self.label_56)
        vlayout.addWidget(self.th_term_line)
        hlayout = QHBoxLayout(self)
        hlayout.addLayout(vlayout)
        hlayout.addWidget(self.textBox)
        vlayout1 = QVBoxLayout(self)
        vlayout1.addLayout(hlayout)
        vlayout1.addWidget(self.add_th)
        self.tabWidget.setTabText(6, "CteateTH")
        self.addTheme.setLayout(vlayout1)

    def create_theme_hours(self):
        data = []
        data.append(self.combo50.currentText())
        data.append(self.combo51.currentText())
        data.append(self.combo52.currentText())
        data.append(self.combo53.currentText())
        data.append(self.combo54.currentText())
        data_id = []
        name_sub = []
        name_gr = []
        name_cour = []
        name_ye = []
        name_org = []
        name_sub.append(data[0])
        name_gr.append(data[1])
        name_cour.append(data[2])
        name_ye.append(data[3])
        name_org.append(data[4])
        sel_sub = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
        sel_gr = 'SELECT `id` FROM `groups` WHERE `name` = %s'
        sel_cour = 'SELECT `id` FROM `courses` WHERE `name` = %s'
        sel_ye = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
        sel_org = 'SELECT `id` FROM `organization` WHERE `name` = %s'
        cursor.execute(sel_sub, name_sub)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_gr, name_gr)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_cour, name_cour)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_ye, name_ye)
        data_id.append(cursor.fetchone()[0])
        cursor.execute(sel_org, name_org)
        data_id.append(cursor.fetchone()[0])
        data_id.append(int(self.th_line.text()))
        data_id.append(int(self.th_term_line.text()))
        data_id.append(self.textBox.toPlainText())
        in_schedule_change = 'INSERT INTO `lessons_plan` (`subjects_id`, `groups_id`, `courses_id`, `year_enter_id`, ' \
                             '`organization_id`, `number`, `term`, `theme`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'
        check_in = 'SELECT * FROM `lessons_plan` WHERE `subjects_id` = %s AND `groups_id` = %s AND ' \
                   '`courses_id` = %s AND `year_enter_id` = %s AND `organization_id` = %s AND `number` = %s AND ' \
                   '`term` = %s AND `theme` = %s'
        cursor.execute(check_in, data_id)
        if cursor.fetchone() == None:
            cursor.execute(in_schedule_change, data_id)
            conn.commit()

    def create_date_type_week_Ui(self):
        self.combo60 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo60.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_60)
        vlayout.addWidget(self.combo60)
        vlayout.addWidget(self.label_61)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.wd_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.add_wd)
        self.tabWidget.setTabText(7, "CteateWD")
        self.addWeekDate.setLayout(vlayout)

    def create_date_type_week(self):
        data = []
        data.append(self.combo60.currentText())
        data_id = []
        name_tw = []
        name_tw.append(data[0])
        sel_tw = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
        cursor.execute(sel_tw, name_tw)
        data_id.append(cursor.fetchone()[0])
        data_id.append(self.wd_date_line.text())
        in_schedule_change = 'INSERT INTO `date_type_week` (`type_week_id`, `date_week`) VALUES (%s, %s)'
        check_in = 'SELECT * FROM `date_type_week` WHERE `type_week_id` = %s AND `date_week` = %s'
        cursor.execute(check_in, data_id)
        if cursor.fetchone() == None:
            cursor.execute(in_schedule_change, data_id)
            conn.commit()

    def update_schedule_Ui(self):
        self.combo30 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo30.addItem(check_sel[i][0])
        self.combo31 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo31.addItem(check_sel[i][0])
        self.combo32 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo32.addItem(check_sel[i][0])
        self.combo33 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo33.addItem(check_sel[i][0])
        self.combo34 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo34.addItem(check_sel[i][0])
        self.combo35 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo35.addItem(check_sel[i][0])
        self.combo36 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `name_day`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo36.addItem(check_sel[i][0])
        self.combo37 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo37.addItem(check_sel[i][0])
        self.combo38 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo38.addItem(check_sel[i][0])
        self.combo39 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `sprav_schedule`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo39.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_71)
        vlayout.addWidget(self.combo30)
        vlayout.addWidget(self.label_73)
        vlayout.addWidget(self.combo31)
        vlayout.addWidget(self.label_74)
        vlayout.addWidget(self.combo32)
        vlayout.addWidget(self.label_75)
        vlayout.addWidget(self.combo33)
        vlayout.addWidget(self.label_76)
        vlayout.addWidget(self.combo34)
        vlayout.addWidget(self.label_77)
        vlayout.addWidget(self.combo35)
        vlayout.addWidget(self.label_78)
        vlayout.addWidget(self.combo36)
        vlayout.addWidget(self.label_79)
        vlayout.addWidget(self.combo37)
        vlayout.addWidget(self.label_72)
        vlayout.addWidget(self.combo38)
        vlayout.addWidget(self.label_80)
        vlayout.addWidget(self.combo39)
        vlayout.addWidget(self.label_81)
        vlayout.addWidget(self.u_schedule_line)
        vlayout.addWidget(self.up_s)
        self.tabWidget.setTabText(8, "UpdateS")
        self.updateSchedule.setLayout(vlayout)

    def update_schedule(self):
        schedule = []
        schedule.append(self.combo30.currentText())
        schedule.append(self.combo31.currentText())
        schedule.append(self.combo32.currentText())
        schedule.append(self.combo33.currentText())
        schedule.append(self.combo34.currentText())
        schedule.append(self.combo35.currentText())
        schedule.append(self.combo36.currentText())
        schedule.append(self.combo37.currentText())
        schedule.append(self.combo38.currentText())
        schedule.append(self.combo39.currentText())
        schedule.append(self.u_schedule_line.text())
        self.US = UpdateSchedule(schedule)
        self.US.show()

    def update_schedule_changes_Ui(self):
        self.combo40 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo40.addItem(check_sel[i][0])
        self.combo41 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo41.addItem(check_sel[i][0])
        self.combo42 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo42.addItem(check_sel[i][0])
        self.combo43 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo43.addItem(check_sel[i][0])
        self.combo44 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo44.addItem(check_sel[i][0])
        self.combo45 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo45.addItem(check_sel[i][0])
        self.combo46 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo46.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_90)
        vlayout.addWidget(self.combo40)
        vlayout.addWidget(self.label_91)
        vlayout.addWidget(self.combo41)
        vlayout.addWidget(self.label_92)
        vlayout.addWidget(self.combo42)
        vlayout.addWidget(self.label_93)
        vlayout.addWidget(self.combo43)
        vlayout.addWidget(self.label_94)
        vlayout.addWidget(self.combo44)
        vlayout.addWidget(self.label_95)
        vlayout.addWidget(self.combo45)
        vlayout.addWidget(self.label_96)
        vlayout.addWidget(self.combo46)
        vlayout.addWidget(self.label_97)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.u_sc_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.label_98)
        vlayout.addWidget(self.u_sc_line)
        vlayout.addWidget(self.up_sc)
        self.tabWidget.setTabText(9, "UpdateSC")
        self.updateScheduleChange.setLayout(vlayout)

    def update_schedule_changes(self):
        schedule_changes = []
        schedule_changes.append(self.combo42.currentText())
        schedule_changes.append(self.combo43.currentText())
        schedule_changes.append(self.combo44.currentText())
        schedule_changes.append(self.combo45.currentText())
        schedule_changes.append(self.combo40.currentText())
        schedule_changes.append(self.combo41.currentText())
        schedule_changes.append(self.u_sc_date_line.text())
        schedule_changes.append(self.combo46.currentText())
        schedule_changes.append(self.u_sc_line.text())
        self.USC = UpdateScheduleChanges(schedule_changes)
        self.USC.show()

    def update_theme_hours_Ui(self):
        self.combo50 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo50.addItem(check_sel[i][0])
        self.combo51 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo51.addItem(check_sel[i][0])
        self.combo52 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo52.addItem(check_sel[i][0])
        self.combo53 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo53.addItem(check_sel[i][0])
        self.combo54 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo54.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_100)
        vlayout.addWidget(self.combo50)
        vlayout.addWidget(self.label_101)
        vlayout.addWidget(self.combo51)
        vlayout.addWidget(self.label_102)
        vlayout.addWidget(self.combo52)
        vlayout.addWidget(self.label_103)
        vlayout.addWidget(self.combo53)
        vlayout.addWidget(self.label_104)
        vlayout.addWidget(self.combo54)
        vlayout.addWidget(self.label_105)
        vlayout.addWidget(self.u_th_line)
        vlayout.addWidget(self.label_106)
        vlayout.addWidget(self.u_th_term_line)
        vlayout.addWidget(self.up_th)
        self.tabWidget.setTabText(10, "UpdateTH")
        self.updateTheme.setLayout(vlayout)

    def update_theme_hours(self):
        theme_hours = []
        theme_hours.append(self.combo50.currentText())
        theme_hours.append(self.combo51.currentText())
        theme_hours.append(self.combo52.currentText())
        theme_hours.append(self.combo53.currentText())
        theme_hours.append(self.combo54.currentText())
        theme_hours.append(self.u_th_line.text())
        theme_hours.append(self.u_th_term_line.text())
        self.UTH = UpdateTheme(theme_hours)
        self.UTH.show()

    def update_date_week_Ui(self):
        self.combo60 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo60.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_110)
        vlayout.addWidget(self.combo60)
        vlayout.addWidget(self.label_111)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.u_wd_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.up_wd)
        self.tabWidget.setTabText(11, "UpdateWD")
        self.updateWeekDate.setLayout(vlayout)

    def update_date_week(self):
        week_date = []
        week_date.append(self.combo60.currentText())
        week_date.append(self.u_wd_date_line.text())
        self.UWD = UpdateWeekDate(week_date)
        self.UWD.show()

    def delete_schedule_Ui(self):
        self.combo30 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo30.addItem(check_sel[i][0])
        self.combo31 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo31.addItem(check_sel[i][0])
        self.combo32 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo32.addItem(check_sel[i][0])
        self.combo33 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo33.addItem(check_sel[i][0])
        self.combo34 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo34.addItem(check_sel[i][0])
        self.combo35 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo35.addItem(check_sel[i][0])
        self.combo36 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `name_day`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo36.addItem(check_sel[i][0])
        self.combo37 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo37.addItem(check_sel[i][0])
        self.combo38 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo38.addItem(check_sel[i][0])
        self.combo39 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `sprav_schedule`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo39.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_121)
        vlayout.addWidget(self.combo30)
        vlayout.addWidget(self.label_123)
        vlayout.addWidget(self.combo31)
        vlayout.addWidget(self.label_124)
        vlayout.addWidget(self.combo32)
        vlayout.addWidget(self.label_125)
        vlayout.addWidget(self.combo33)
        vlayout.addWidget(self.label_126)
        vlayout.addWidget(self.combo34)
        vlayout.addWidget(self.label_127)
        vlayout.addWidget(self.combo35)
        vlayout.addWidget(self.label_128)
        vlayout.addWidget(self.combo36)
        vlayout.addWidget(self.label_129)
        vlayout.addWidget(self.combo37)
        vlayout.addWidget(self.label_122)
        vlayout.addWidget(self.combo38)
        vlayout.addWidget(self.label_130)
        vlayout.addWidget(self.combo39)
        vlayout.addWidget(self.label_131)
        vlayout.addWidget(self.d_schedule_line)
        vlayout.addWidget(self.del_s)
        self.tabWidget.setTabText(12, "DeleteS")
        self.deleteSchedule.setLayout(vlayout)

    def delete_schedule(self):
        schedule = []
        schedule.append(self.combo30.currentText())
        schedule.append(self.combo31.currentText())
        schedule.append(self.combo32.currentText())
        schedule.append(self.combo33.currentText())
        schedule.append(self.combo34.currentText())
        schedule.append(self.combo35.currentText())
        schedule.append(self.combo36.currentText())
        schedule.append(self.combo37.currentText())
        schedule.append(self.combo38.currentText())
        schedule.append(self.combo39.currentText())
        schedule.append(self.d_schedule_line.text())
        self.DS = DeleteSchedule(schedule)
        self.DS.show()

    def delete_schedule_changes_Ui(self):
        self.combo40 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo40.addItem(check_sel[i][0])
        self.combo41 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `teachers`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo41.addItem(check_sel[i][0])
        self.combo42 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo42.addItem(check_sel[i][0])
        self.combo43 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo43.addItem(check_sel[i][0])
        self.combo44 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo44.addItem(check_sel[i][0])
        self.combo45 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `num_lessons`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo45.addItem(check_sel[i][0])
        self.combo46 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo46.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_140)
        vlayout.addWidget(self.combo40)
        vlayout.addWidget(self.label_141)
        vlayout.addWidget(self.combo41)
        vlayout.addWidget(self.label_142)
        vlayout.addWidget(self.combo42)
        vlayout.addWidget(self.label_143)
        vlayout.addWidget(self.combo43)
        vlayout.addWidget(self.label_144)
        vlayout.addWidget(self.combo44)
        vlayout.addWidget(self.label_145)
        vlayout.addWidget(self.combo45)
        vlayout.addWidget(self.label_146)
        vlayout.addWidget(self.combo46)
        vlayout.addWidget(self.label_147)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.d_sc_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.label_148)
        vlayout.addWidget(self.d_sc_line)
        vlayout.addWidget(self.del_sc)
        self.tabWidget.setTabText(13, "DeleteSC")
        self.deleteScheduleChange.setLayout(vlayout)

    def delete_schedule_changes(self):
        schedule_changes = []
        schedule_changes.append(self.combo42.currentText())
        schedule_changes.append(self.combo43.currentText())
        schedule_changes.append(self.combo44.currentText())
        schedule_changes.append(self.combo45.currentText())
        schedule_changes.append(self.combo40.currentText())
        schedule_changes.append(self.combo41.currentText())
        schedule_changes.append(self.d_sc_date_line.text())
        schedule_changes.append(self.combo46.currentText())
        schedule_changes.append(self.d_sc_line.text())
        self.DSC = DeleteScheduleChanges(schedule_changes)
        self.DSC.show()

    def delete_theme_hours_Ui(self):
        self.combo50 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `subjects`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo50.addItem(check_sel[i][0])
        self.combo51 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `groups`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo51.addItem(check_sel[i][0])
        self.combo52 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `courses`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo52.addItem(check_sel[i][0])
        self.combo53 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `year_enter`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo53.addItem(check_sel[i][0])
        self.combo54 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `organization`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo54.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_150)
        vlayout.addWidget(self.combo50)
        vlayout.addWidget(self.label_151)
        vlayout.addWidget(self.combo51)
        vlayout.addWidget(self.label_152)
        vlayout.addWidget(self.combo52)
        vlayout.addWidget(self.label_153)
        vlayout.addWidget(self.combo53)
        vlayout.addWidget(self.label_154)
        vlayout.addWidget(self.combo54)
        vlayout.addWidget(self.label_155)
        vlayout.addWidget(self.d_th_line)
        vlayout.addWidget(self.label_156)
        vlayout.addWidget(self.d_th_term_line)
        vlayout.addWidget(self.del_th)
        self.tabWidget.setTabText(14, "DeleteTH")
        self.deleteTheme.setLayout(vlayout)

    def delete_theme_hours(self):
        theme_hours = []
        theme_hours.append(self.combo50.currentText())
        theme_hours.append(self.combo51.currentText())
        theme_hours.append(self.combo52.currentText())
        theme_hours.append(self.combo53.currentText())
        theme_hours.append(self.combo54.currentText())
        theme_hours.append(self.d_th_line.text())
        theme_hours.append(self.d_th_term_line.text())
        self.DTH = DeleteThemeHours(theme_hours)
        self.DTH.show()

    def delete_date_week_Ui(self):
        self.combo60 = QComboBox(self)
        cursor.execute('SELECT `name` FROM `type_week`')
        check_sel = cursor.fetchall()
        for i in range(0, len(check_sel)):
            self.combo60.addItem(check_sel[i][0])
        self.tabWidget = QTabWidget(self.centralwidget)
        vlayout = QVBoxLayout(self)
        vlayout.addWidget(self.label_160)
        vlayout.addWidget(self.combo60)
        vlayout.addWidget(self.label_161)
        hlayout = QHBoxLayout(self)
        hlayout.addWidget(self.d_wd_date_line)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.del_wd)
        self.tabWidget.setTabText(15, "DeleteWD")
        self.deleteWeekDate.setLayout(vlayout)

    def delete_date_week(self):
        date_week = []
        date_week.append(self.combo60.currentText())
        date_week.append(self.d_wd_date_line.text())
        self.DWD = DeleteWeekDate(date_week)
        self.DWD.show()

    def parser_th_Ui(self):
        cursor.execute('SELECT `name` FROM `organization`')
        name_org = cursor.fetchall()
        self.combo73 = QComboBox(self)
        for i in range(0, len(name_org)):
            self.combo73.addItem(name_org[i][0])
        vl = QVBoxLayout(self)
        vl.addWidget(self.combo73)
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label_3)
        hlayout.addWidget(self.filename_3)
        hlayout.addWidget(self.open_pars_th)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(vl)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_th)
        self.tabWidget.setTabText(3, "ParserTH")
        self.parser_th.setLayout(vlayout)

    def parser_theme_h(self):
        df = pd.read_excel(io=self.file_, engine='openpyxl', sheet_name='Лист1')
        arr = list(df.head(0))

        # Парс excel
        result = []

        # Парс ГРУПП
        for i in range(0, len(df[arr[2]].tolist())):
            add_ser = 'INSERT INTO `groups` (`name`) VALUES (%s)'
            result.append(df[arr[2]].tolist()[i])
            check_input = 'SELECT * FROM `groups` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс Курс
        for i in range(0, len(df[arr[3]].tolist())):
            add_ser = 'INSERT INTO `courses` (`name`) VALUES (%s)'
            result.append(df[arr[3]].tolist()[i])
            check_input = 'SELECT * FROM `courses` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГОДА ПОСТУАЛЕНИЯ
        for i in range(0, len(df[arr[4]].tolist())):
            add_ser = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
            result.append(df[arr[4]].tolist()[i])
            check_input = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ЧАСОВ ДЛЯ ТЕМ
        arr = list(df.head(0))
        for i in range(0, len(df[arr[0]].tolist())):
            add_ser = 'INSERT INTO `lessons_plan` (`subjects_id`, `theme`, `groups_id`, `courses_id`, `year_enter_id`, ' \
                      '`number`, `organization_id`, `term`) ' \
                      'VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'
            for j in range(0, len(arr)):
                data_db = []
                if j == 0:
                    data_db.append(df.values.tolist()[i][j])
                    check_input = 'SELECT `id` FROM `subjects` WHERE `name` LIKE %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                elif j == 1:
                    result.append(df.values.tolist()[i][j])
                elif j == 2:
                    data_db.append(df.values.tolist()[i][j])
                    check_input = 'SELECT `id` FROM `groups` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                elif j == 3:
                    data_db.append(df.values.tolist()[i][j])
                    check_input = 'SELECT `id` FROM `courses` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                elif j == 4:
                    data_db.append(df.values.tolist()[i][j])
                    check_input = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                elif j == 5:
                    result.append(int(df.values.tolist()[i][j]))
                elif j == 6:
                    data_db.append(self.combo73.currentText())
                    check_input = 'SELECT `id` FROM `organization` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    result.append(df.values.tolist()[i][j])
            check_input = 'SELECT * FROM `lessons_plan` WHERE `subjects_id` = %s AND `theme` = %s AND `groups_id` = %s AND ' \
                          '`courses_id` = %s AND `year_enter_id` = %s AND `number` = %s AND `organization_id` = %s AND ' \
                          '`term` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

    def pars_win_th(self):
        self.open_pars_th.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename_3.setText(fname)
            self.file_ = str(fname)

    def parser_date_week_Ui(self):
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label_2)
        hlayout.addWidget(self.filename_2)
        hlayout.addWidget(self.open_pars_date_week)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_date_week)
        self.tabWidget.setTabText(0, "Parser")
        self.parser_date_week_.setLayout(vlayout)

    def parser_date_week(self):
        df = pd.read_excel(io=self.file_, engine='openpyxl', sheet_name='Лист1')
        arr = list(df.head(0))

        # Парс excel
        result = []

        # Парс ДАТ НЕДЕЛЬ
        for n in range(0, len(arr)):
            for i in range(0, len(df[arr[n]].tolist())):
                dates = df[arr[n]].tolist()[i]
                str_start_date = ''
                str_end_date = ''
                for j in range(0, len(dates)):
                    if dates[j] != '-' and j < 10:
                        str_start_date += dates[j]
                    elif dates[j] != '-' and j > 10:
                        str_end_date += dates[j]
                day = ''
                mouth = ''
                year = ''
                for j in range(0, len(str_start_date)):
                    if j < 2 and str_start_date[j] != '.':
                        day += str_start_date[j]
                    elif j > 2 and j < 5 and str_start_date[j] != '.':
                        mouth += str_start_date[j]
                    elif j > 5 and str_start_date[j] != '.':
                        year += str_start_date[j]
                start_date = datetime(int(year), int(mouth), int(day))
                day = ''
                mouth = ''
                year = ''
                for j in range(0, len(str_end_date)):
                    if j < 2 and str_end_date[j] != '.':
                        day += str_end_date[j]
                    elif j > 2 and j < 5 and str_end_date[j] != '.':
                        mouth += str_end_date[j]
                    elif j > 5 and str_end_date[j] != '.':
                        year += str_end_date[j]
                end_date = datetime(int(year), int(mouth), int(day))
                res = pd.date_range(min(start_date, end_date), max(start_date, end_date)).strftime('%Y-%m-%d').tolist()
                for j in range(0, len(res)):
                    cursor.execute("SELECT `id` FROM `type_week` WHERE `name` LIKE %s", [arr[n]])
                    result.append(cursor.fetchone()[0])
                    result.append(res[j])
                    add_res = 'INSERT INTO `date_type_week` (`type_week_id`, `date_week`) VALUES (%s, %s)'
                    cursor.execute(add_res, result)
                    conn.commit()
                    result = []

    def pars_win_date_week(self):
        self.open_pars_date_week.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename_2.setText(fname)
            self.file_ = str(fname)

    def parser_schedule(self):

        df = pd.read_excel(io=self.file_s, engine='openpyxl', sheet_name='Лист1')
        excel = openpyxl.load_workbook(filename=self.file_s)
        sheet = excel.worksheets[0]

        # Парс excel
        result = []

        # Парс НАЗВАНИЯ РАСПИСНАИЯ
        add_ser = 'INSERT INTO `sprav_schedule` (`name`) VALUES (%s)'
        result.append(self.file_s)
        check_input = 'SELECT * FROM `sprav_schedule` WHERE `name` = %s'
        cursor.execute(check_input, result)
        self.path_line.setText(result[0])
        if cursor.fetchone() == None:
            cursor.execute(add_ser, result)
            conn.commit()
            result = []
        else:
            result = []

        for r in sheet.merged_cells.ranges:
            cl, rl, cr, rr = r.bounds  # границы объединенной области
            rl -= 2
            rr -= 1
            cl -= 1
            base_value = df.iloc[rl, cl]
            df.iloc[rl:rr, cl:cr] = base_value

        # Парс ДНЕЙ
        for i in range(0, len(df['Дни'].tolist())):
            add_ser = 'INSERT INTO `name_day` (`name`) VALUES (%s)'
            result.append(df['Дни'].tolist()[i])
            check_input = 'SELECT * FROM `name_day` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ПАР
        for i in range(0, len(df['Уроки'].tolist())):
            add_ser = 'INSERT INTO `num_lessons` (`name`) VALUES (%s)'
            result.append(df['Уроки'].tolist()[i])
            check_input = 'SELECT * FROM `num_lessons` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ТИПА НЕДЕЛИ
        for i in range(0, len(df['Unnamed: 2'].tolist())):
            add_ser = 'INSERT INTO `type_week` (`name`) VALUES (%s)'
            result.append(df['Unnamed: 2'].tolist()[i])
            check_input = 'SELECT * FROM `type_week` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГРУПП ПЕРВОГО КУРСА
        res = []
        arr = list(df.head(0))
        gr = ''
        for j in range(0, len(arr[3])):
            if j < arr[3].find('-'):
                if arr[3][j] != ',' and arr[3][j].isupper():
                    gr += arr[3][j]
                else:
                    if len(gr) > 0:
                        res.append(gr)
                        gr = ''
            elif len(gr) > 0:
                res.append(gr)
                gr = ''
        for i in range(0, len(res)):
            result.append(res[i])
            add_ser = 'INSERT INTO `groups` (`name`) VALUES (%s)'
            check_input = 'SELECT * FROM `groups` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГОДА ПОСТУПЛЕНИЯ ПЕРВЫХ КУРСОВ
        arr = list(df.head(0))
        ye = ''
        for j in range(0, len(arr[3])):
            if j > arr[3].find('-'):
                if arr[3][j] != '-':
                    ye += arr[3][j]
        result.append(ye)
        add_ser = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
        check_input = 'SELECT * FROM `year_enter` WHERE `name` = %s'
        cursor.execute(check_input, result)
        if cursor.fetchone() == None:
            cursor.execute(add_ser, result)
            conn.commit()
            result = []
        else:
            result = []

        # Парс ДИСЦИЛИН И ПРЕПОДААВТЕЛЕЙ
        dis = ''
        tea = ''
        arr = list(df.head(0))
        for n in range(3, len(arr)):
            for j in range(0, len(df[arr[n]].tolist())):
                if isinstance(df[arr[n]][j], str):
                    for i in range(0, len(df[arr[n]][j])):
                        if i < df[arr[n]][j].find('\n'):
                            dis += df[arr[n]][j][i]
                        elif i > df[arr[n]][j].find('\n'):
                            if df[arr[n]][j][i] == ' ' and df[arr[n]][j][i + 1] == ' ':
                                break
                            else:
                                tea += df[arr[n]][j][i]
                    right = False
                    dis_arr = []
                    for i in range(0, len(dis)):
                        dis_arr.append(dis[i])
                    dis = ''
                    while right != True:
                        if dis_arr[-1:][0] == ' ':
                            dis_arr.pop(-1)
                        else:
                            right = True
                    for i in range(0, len(dis_arr)):
                        dis += dis_arr[i]
                    result.append(dis)
                    add_ser = 'INSERT INTO `subjects` (`name`) VALUES (%s)'
                    check_input = 'SELECT * FROM `subjects` WHERE `name` LIKE %s'
                    cursor.execute(check_input, result)
                    if cursor.fetchone() == None:
                        cursor.execute(add_ser, result)
                        conn.commit()
                        result = []
                        dis = ''
                    else:
                        result = []
                        dis = ''
                    result.append(tea)
                    add_ser = 'INSERT INTO `teachers` (`name`) VALUES (%s)'
                    check_input = 'SELECT * FROM `teachers` WHERE `name` = %s'
                    cursor.execute(check_input, result)
                    if cursor.fetchone() == None:
                        cursor.execute(add_ser, result)
                        conn.commit()
                        result = []
                        tea = ''
                    else:
                        result = []
                        tea = ''

        # Парс ВСЕХ ГРУПП КРОМЕ ЮРИСТОВ
        arr = list(df.head(0))
        res = []
        gr = ''
        for i in range(3, len(arr)):
            if len(arr[i]) <= 9:
                for j in range(0, len(arr[i])):
                    if j < arr[i].find('-'):
                        if arr[i][j].isupper():
                            gr += arr[i][j]
                res.append(gr)
                gr = ''
        for i in range(0, len(res)):
            result.append(res[i])
            add_ser = 'INSERT INTO `groups` (`name`) VALUES (%s)'
            check_input = 'SELECT * FROM `groups` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ВСЕХ ГРУПП ЮРИСТОВ
        arr = list(df.head(0))
        res = []
        gr = ''
        for i in range(4, len(arr)):
            if len(arr[i]) > 9:
                for j in range(0, len(arr[i])):
                    if j < arr[i].find('-'):
                        if arr[i][j].isupper():
                            gr += arr[i][j]
                res.append(gr)
                gr = ''
        for i in range(0, len(res)):
            result.append(res[i])
            add_ser = 'INSERT INTO `groups` (`name`) VALUES (%s)'
            check_input = 'SELECT * FROM `groups` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс ГОДА ПОСТПЛЕНИЕ ВСЕХ КУРСОВ КРОМЕ ПЕРВОГО
        arr = list(df.head(0))
        ye = ''
        res = []
        for i in range(4, len(arr)):
            for j in range(0, len(arr[i])):
                if j > arr[i].find('-'):
                    if j < arr[i].find('/'):
                        ye += arr[i][j]
                    if len(arr[i]) <= 9:
                        ye += arr[i][j]
            res.append(ye)
            ye = ''
        for i in range(0, len(res)):
            result.append(res[i])
            add_ser = 'INSERT INTO `year_enter` (`name`) VALUES (%s)'
            check_input = 'SELECT * FROM `year_enter` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс КУРСОВ
        arr = list(df.head(0))
        ye = ''
        res = []
        current_y = datetime.now().strftime('%Y')
        current_m = datetime.now().strftime('%m')
        for i in range(4, len(arr)):
            for j in range(0, len(arr[i])):
                if j > arr[i].find('-'):
                    if j < arr[i].find('/'):
                        ye += arr[i][j]
                    if len(arr[i]) <= 9:
                        ye += arr[i][j]
            res.append(ye)
            ye = ''
        for i in range(0, len(res)):
            if int(current_m) >= 9:
                result.append((int(current_y) - int(res[i])) + 1)
            else:
                result.append(int(current_y) - int(res[i]))
            add_ser = 'INSERT INTO `courses` (`name`) VALUES (%s)'
            check_input = 'SELECT * FROM `courses` WHERE `name` = %s'
            cursor.execute(check_input, result)
            if cursor.fetchone() == None:
                cursor.execute(add_ser, result)
                conn.commit()
                result = []
            else:
                result = []

        # Парс РАСПИСАНИЯ ДЛЯ ПЕРВОГО КУРСА
        res_g = []
        arr = list(df.head(0))
        gr = ''
        for j in range(0, len(arr[3])):
            if j < arr[3].find('-'):
                if arr[3][j] != ',' and arr[3][j].isupper():
                    gr += arr[3][j]
                else:
                    if len(gr) > 0:
                        res_g.append(gr)
                        gr = ''
            elif len(gr) > 0:
                res_g.append(gr)
                gr = ''
        ye = ''
        for j in range(0, len(arr[3])):
            if j > arr[3].find('-'):
                if arr[3][j] != '-':
                    ye += arr[3][j]
        dis = ''
        tea = ''
        res_d = []
        res_t = []
        for j in range(0, len(df[arr[3]].tolist())):
            if isinstance(df[arr[3]][j], str):
                for i in range(0, len(df[arr[3]][j])):
                    if i < df[arr[3]][j].find('\n'):
                        dis += df[arr[3]][j][i]
                    elif i > df[arr[3]][j].find('\n'):
                        if df[arr[3]][j][i] == ' ' and df[arr[3]][j][i + 1] == ' ':
                            break
                        else:
                            tea += df[arr[3]][j][i]
                right = False
                dis_arr = []
                for i in range(0, len(dis)):
                    dis_arr.append(dis[i])
                dis = ''
                while right != True:
                    if dis_arr[-1:][0] == ' ':
                        dis_arr.pop(-1)
                    else:
                        right = True
                for i in range(0, len(dis_arr)):
                    dis += dis_arr[i]
                res_d.append(dis)
                res_t.append(tea)
                dis = ''
                tea = ''
            else:
                res_d.append(None)
                res_t.append(None)
                dis = ''
                tea = ''
        current_y = datetime.now().strftime('%Y')
        current_m = datetime.now().strftime('%m')
        if int(current_m) >= 9:
            res_c = (int(current_y) - int(ye)) + 1
        else:
            res_c = int(current_y) - int(ye)
        for i in range(0, len(df[arr[3]].tolist())):
            add_ser = 'INSERT INTO `schedule` (`name_day_id`, `num_lessons_id`, `type_week_id`, `groups_id`, ' \
                      '`year_enter_id`, `subjects_id`, `teachers_id`, `courses_id`, num_group, `organization_id`, ' \
                      '`sprav_schedule_id`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            for n in range(0, len(res_g)):
                for j in range(0, len(arr)):
                    data_db = []
                    if j == 0:
                        data_db.append(df[arr[j]][i])
                        check_input = 'SELECT `id` FROM `name_day` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    if j == 1:
                        data_db.append(df[arr[j]][i])
                        check_input = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    if j == 2:
                        data_db.append(df[arr[j]][i])
                        check_input = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    if j == 3:
                        data_db.append(res_g[n])
                        check_input = 'SELECT `id` FROM `groups` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                        data_db = []
                        data_db.append(ye)
                        check_input = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                        data_db = []
                        data_db.append(res_d[i])
                        if res_d[i] != None:
                            check_input = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
                            cursor.execute(check_input, data_db)
                            result.append(cursor.fetchone()[0])
                        else:
                            result.append(None)
                        data_db = []
                        data_db.append(res_t[i])
                        if res_t[i] != None:
                            check_input = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
                            cursor.execute(check_input, data_db)
                            result.append(cursor.fetchone()[0])
                        else:
                            result.append(None)
                        data_db = []
                        data_db.append(res_c)
                        check_input = 'SELECT `id` FROM `courses` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                        if arr[3].find('/') < 0:
                            result.append(1)
                        else:
                            for s in range(0, len(arr[3])):
                                if s > arr[3].find('/'):
                                    result.append(int(arr[3][s]))
                        data_db = []
                        data_db.append(self.combo70.currentText())
                        check_input = 'SELECT `id` FROM `organization` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                        data_db = []
                        data_db.append(self.file_s)
                        check_input = 'SELECT `id` FROM `sprav_schedule` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                check_input = 'SELECT * FROM `schedule` WHERE `name_day_id` = %s AND `num_lessons_id` = %s AND ' \
                              '`type_week_id` = %s AND `groups_id` = %s AND `year_enter_id` = %s AND `subjects_id` = %s ' \
                              'AND `teachers_id` = %s AND `courses_id` = %s AND num_group = %s AND `organization_id` = %s ' \
                              'AND `sprav_schedule_id` = %s'
                cursor.execute(check_input, result)
                if cursor.fetchone() == None:
                    cursor.execute(add_ser, result)
                    conn.commit()
                    result = []
                else:
                    result = []

        # Парс РАСПИСАНИЯ
        arr = list(df.head(0))
        res_g = []
        gr = ''
        res_g.append('')
        res_g.append('')
        res_g.append('')
        res_g.append('')
        for i in range(4, len(arr)):
            if len(arr[i]) <= 9:
                for j in range(0, len(arr[i])):
                    if j < arr[i].find('-'):
                        if arr[i][j].isupper():
                            gr += arr[i][j]
            elif len(arr[i]) > 9:
                for j in range(0, len(arr[i])):
                    if j < arr[i].find('-'):
                        if arr[i][j].isupper():
                            gr += arr[i][j]
            res_g.append(gr)
            gr = ''
        ye = ''
        res_ye = []
        res_ye.append('')
        res_ye.append('')
        res_ye.append('')
        res_ye.append('')
        for i in range(4, len(arr)):
            for j in range(0, len(arr[i])):
                if j > arr[i].find('-'):
                    if j < arr[i].find('/'):
                        ye += arr[i][j]
                    if len(arr[i]) <= 9:
                        ye += arr[i][j]
            res_ye.append(ye)
            ye = ''
        current_y = datetime.now().strftime('%Y')
        current_m = datetime.now().strftime('%m')
        for i in range(0, len(df[arr[3]].tolist())):
            add_ser = 'INSERT INTO `schedule` (`name_day_id`, `num_lessons_id`, `type_week_id`, `groups_id`, ' \
                      '`year_enter_id`, `subjects_id`, `teachers_id`, `courses_id`, num_group, `organization_id`, `sprav_schedule_id`) ' \
                      'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            data_nd = 0
            data_nl = 0
            data_tw = 0
            for j in range(0, len(arr)):
                data_db = []
                dis = ''
                tea = ''
                if j == 0:
                    data_db.append(df[arr[j]][i])
                    check_input = 'SELECT `id` FROM `name_day` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    data_nd = cursor.fetchone()[0]
                if j == 1:
                    data_db.append(df[arr[j]][i])
                    check_input = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    data_nl = cursor.fetchone()[0]
                if j == 2:
                    data_db.append(df[arr[j]][i])
                    check_input = 'SELECT `id` FROM `type_week` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    data_tw = cursor.fetchone()[0]
                if j >= 4:
                    result.append(data_nd)
                    result.append(data_nl)
                    result.append(data_tw)
                    data_db.append(res_g[j])
                    check_input = 'SELECT `id` FROM `groups` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    data_db = []
                    data_db.append(res_ye[j])
                    check_input = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    data_db = []
                    if isinstance(df[arr[j]][i], str):
                        for n in range(0, len(df[arr[j]][i])):
                            if n < df[arr[j]][i].find('\n'):
                                dis += df[arr[j]][i][n]
                            elif n > df[arr[j]][i].find('\n'):
                                if df[arr[j]][i][n] == ' ' and df[arr[j]][i][n + 1] == ' ':
                                    break
                                else:
                                    tea += df[arr[j]][i][n]
                        right = False
                        dis_arr = []
                        for s in range(0, len(dis)):
                            dis_arr.append(dis[s])
                        dis = ''
                        while right != True:
                            if dis_arr[-1:][0] == ' ':
                                dis_arr.pop(-1)
                            else:
                                right = True
                        for s in range(0, len(dis_arr)):
                            dis += dis_arr[s]
                    else:
                        dis = None
                        tea = None
                    data_db.append(dis)
                    if dis != None:
                        check_input = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    else:
                        result.append(None)
                    data_db = []
                    data_db.append(tea)
                    if tea != None:
                        check_input = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    else:
                        result.append(None)
                    data_db = []
                    if int(current_m) >= 9:
                        res_c = (int(current_y) - int(res_ye[j])) + 1
                    else:
                        res_c = int(current_y) - int(res_ye[j])
                    data_db.append(res_c)
                    check_input = 'SELECT `id` FROM `courses` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    if arr[j].find('/') < 0:
                        result.append(1)
                    else:
                        for s in range(0, len(arr[j])):
                            if s > arr[j].find('/'):
                                result.append(int(arr[j][s]))
                                break
                    data_db = []
                    data_db.append(self.combo70.currentText())
                    check_input = 'SELECT `id` FROM `organization` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    data_db = []
                    data_db.append(self.file_s)
                    check_input = 'SELECT `id` FROM `sprav_schedule` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    check_input = 'SELECT * FROM `schedule` WHERE `name_day_id` = %s AND `num_lessons_id` = %s AND ' \
                                  '`type_week_id` = %s AND `groups_id` = %s AND `year_enter_id` = %s AND `subjects_id` = %s ' \
                                  'AND `teachers_id` = %s AND `courses_id` = %s AND num_group = %s AND `organization_id` = %s ' \
                                  'AND `sprav_schedule_id` = %s'
                    cursor.execute(check_input, result)
                    if cursor.fetchone() == None:
                        cursor.execute(add_ser, result)
                        conn.commit()
                        result = []
                    else:
                        result = []
        self.dlg = QMessageBox()
        self.dlg.addButton("Да", QMessageBox.AcceptRole)
        self.dlg.addButton("Нет", QMessageBox.AcceptRole)
        self.dlg.setIcon(QMessageBox.Information)
        self.dlg.setWindowTitle("Действие")
        self.dlg.setInformativeText(
            "Сзделать загруженное расписание основным?")
        bttn = self.dlg.exec()
        if self.dlg.clickedButton().text() == "Да":
            self.save_base_schedule()

    def check_save_base_schedule(self):
        return os.path.exists('config_base_schedule.json')

    def save_base_schedule(self):
        if self.check_save_base_schedule():
            data = [{'path_line': self.path_line.text()}]
            with open('config_base_schedule.json', 'w') as save:
                json.dump(data, save)
        else:
            workbook = Workbook()
            worksheet = workbook.worksheets[0]
            worksheet.cells.get("A1").put_value("path")
            worksheet.cells.get("A2").put_value(self.path_line.text())
            workbook.save("config_base_schedule.json")

    def parser_schedule_Ui(self):
        cursor.execute('SELECT `name` FROM `organization`')
        name_org = cursor.fetchall()
        self.combo70 = QComboBox(self)
        for i in range(0, len(name_org)):
            self.combo70.addItem(name_org[i][0])
        vl = QVBoxLayout(self)
        vl.addWidget(self.combo70)
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label)
        hlayout.addWidget(self.filename)
        hlayout.addWidget(self.open_pars_s)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(vl)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_s)
        self.tabWidget.setTabText(1, "ParserS")
        self.parser_s.setLayout(vlayout)

    def parser_schedule_changes(self):

        df = pd.read_excel(io=self.file_sc, engine='openpyxl', sheet_name='Лист1')
        excel = openpyxl.load_workbook(filename=self.file_sc)
        sheet = excel.worksheets[0]
        slovar = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '.']

        # Парс excel
        result = []

        # Парс ЗАМЕН
        arr = list(df.head(0))
        date_changes = ''
        for i in range(0, len(arr[0])):
            for n in range(0, len(slovar)):
                if arr[0][i] == slovar[n]:
                    date_changes += arr[0][i]
        d = ''
        m = ''
        y = ''
        step = 0
        for i in range(0, len(date_changes)):
            if date_changes[i] != '.' and step == 0:
                d += date_changes[i]
            if date_changes[i] == '.':
                step += 1
            if date_changes[i] != '.' and step == 1:
                m += date_changes[i]
            if date_changes[i] != '.' and step == 2:
                y += date_changes[i]
        date_changes = date(int(y), int(m), int(d))
        time_format = "%Y-%m-%d"
        for i in range(1, len(df[arr[0]].tolist())):
            nl = []
            add_ser = 'INSERT INTO `schedule_changes` (`groups_id`, `courses_id`, `year_enter_id`, `num_lessons_id`, ' \
                      '`teachers_id`, `subjects_id`, `date_changes`, `organization_id`, `num_group`) VALUES (' \
                      '%s, %s, %s, %s, %s, %s, %s, %s, %s)'
            for j in range(0, len(arr[0])):
                data_db = []
                if j == 0:
                    gr = ''
                    course = ''
                    for n in range(0, len(df[arr[j]][i])):
                        if n > df[arr[j]][i].find('.'):
                            gr += df[arr[j]][i][n]
                        elif n < df[arr[j]][i].find('.') and df[arr[j]][i][n] != 'к':
                            course += df[arr[j]][i][n]
                    current_y = datetime.now().strftime('%Y')
                    current_m = datetime.now().strftime('%m')
                    if int(current_m) >= 9:
                        res_ye = (int(current_y) - int(course)) + 1
                    else:
                        res_ye = int(current_y) - int(course)
                    if gr == 'ЮР':
                        gr = 'ПСО'
                    data_db.append(gr)
                    check_input = 'SELECT `id` FROM `groups` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    data_db = []
                    data_db.append(course)
                    check_input = 'SELECT `id` FROM `courses` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    data_db = []
                    data_db.append(res_ye)
                    check_input = 'SELECT `id` FROM `year_enter` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                if j == 1:
                    if str(df[arr[j]][i]).find(',') > 0:
                        for s in range(0, len(df[arr[j]][i])):
                            if s < str(df[arr[j]][i]).find(',') or s > str(df[arr[j]][i]).find(',') and str(df[arr[j]][i][s]) != ' ':
                                nl.append(str(df[arr[j]][i][s]))
                    else:
                        data_db.append(str(df[arr[j]][i])+'.0')
                        check_input = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                if j == 5:
                    if df[arr[j]][i] == '-':
                        result.append(None)
                    else:
                        data_db.append(str(df[arr[j]][i]))
                        check_input = 'SELECT `id` FROM `teachers` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                if j == 6:
                    if df[arr[j]][i] == '-':
                        result.append(None)
                    else:
                        data_db.append(str(df[arr[j]][i]))
                        check_input = 'SELECT `id` FROM `subjects` WHERE `name` = %s'
                        cursor.execute(check_input, data_db)
                        result.append(cursor.fetchone()[0])
                    result.append(f"{date_changes:{time_format}}")
                    data_db = []
                    data_db.append(self.combo71.currentText())
                    check_input = 'SELECT `id` FROM `organization` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    if df[arr[j]][i].find('/') > 0:
                        for s in range(0, len(df[arr[j]][i])):
                            if s > df[arr[j]][i].find('/'):
                                result.append(int(df[arr[j]][i][s]))
                                break
                    else:
                        result.append(1)
            if len(nl) > 0:
                result_old = list(result)
                for j in range(0, len(nl)):
                    result = list(result_old)
                    data_db = []
                    data_db.append(str(nl[j]) + '.0')
                    check_input = 'SELECT `id` FROM `num_lessons` WHERE `name` = %s'
                    cursor.execute(check_input, data_db)
                    result.append(cursor.fetchone()[0])
                    add_ser = 'INSERT INTO `schedule_changes` (`groups_id`, `courses_id`, `year_enter_id`, ' \
                              '`teachers_id`, `subjects_id`, `date_changes`, `organization_id`, `num_group`, `num_lessons_id`) VALUES (' \
                              '%s, %s, %s, %s, %s, %s, %s, %s, %s)'
                    check_input = 'SELECT * FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                                  '`year_enter_id` = %s AND `teachers_id` = %s AND `subjects_id` = %s AND ' \
                                  '`date_changes` = %s AND `organization_id` = %s AND `num_group` = %s AND ' \
                                  '`num_lessons_id` = %s'
                    cursor.execute(check_input, result)
                    if cursor.fetchone() == None:
                        cursor.execute(add_ser, result)
                        conn.commit()
                        result = []
                    else:
                        result = []
            else:
                check_input = 'SELECT * FROM `schedule_changes` WHERE `groups_id` = %s AND `courses_id` = %s AND ' \
                              '`year_enter_id` = %s AND `num_lessons_id` = %s AND `teachers_id` = %s AND `subjects_id` ' \
                              '= %s AND `date_changes` = %s AND `organization_id` = %s AND `num_group` = %s'
                cursor.execute(check_input, result)
                if cursor.fetchone() == None:
                    cursor.execute(add_ser, result)
                    conn.commit()
                    result = []
                else:
                    result = []

    def parser_schedule_changes_Ui(self):
        cursor.execute('SELECT `name` FROM `organization`')
        name_org = cursor.fetchall()
        self.combo71 = QComboBox(self)
        for i in range(0, len(name_org)):
            self.combo71.addItem(name_org[i][0])
        vl = QVBoxLayout(self)
        vl.addWidget(self.combo71)
        hlayout = QHBoxLayout(self)
        self.tabWidget = QTabWidget(self.centralwidget)
        hlayout.addWidget(self.label1)
        hlayout.addWidget(self.filename1)
        hlayout.addWidget(self.open_pars_sc)
        vlayout = QVBoxLayout(self)
        vlayout.addLayout(vl)
        vlayout.addLayout(hlayout)
        vlayout.addWidget(self.pars_sc)
        self.tabWidget.setTabText(2, "ParserSC")
        self.parser_sc.setLayout(vlayout)

    def pars_win_schedule_changes(self):
        self.open_pars_sc.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename1.setText(fname)
            self.file_sc = str(fname)

    def pars_win_schedule(self):
        self.open_pars_s.hide()
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', './', '(*.xls *.xlsx)')
        if fname:
            self.filename.setText(fname)
            self.file_s = str(fname)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    splash = QtWidgets.QSplashScreen()
    splash.setPixmap(QtGui.QPixmap('images/splash.jpg'))
    splash.show()
    splash.showMessage('<h1 style="color:#ffffff;">Добро пожаловать в АИС Деканат (beta)</h1>',
                       QtCore.Qt.AlignTop | QtCore.Qt.AlignLeft, QtCore.Qt.white)
    QtCore.QThread.msleep(5000)
    w = Login()
    w.show()
    splash.hide()
    sys.exit(app.exec_())
